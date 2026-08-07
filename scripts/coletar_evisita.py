# ============================================================
# ANÁLISE DE TEMPO DE VISITAS - E-VISITA (SES/MS)
# Coleta pendência "Não" + tipo_visita=0 (normal + recuperação)
# Nome após todas páginas + Duplicadas + Horários por Turno
# ============================================================

import os
import csv
import re
import json
import shutil
import time
import unicodedata
import logging
import argparse
from itertools import groupby
from datetime import date, datetime, timedelta, time as dt_time
from concurrent.futures import ThreadPoolExecutor, as_completed

import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

import pandas as pd
from bs4 import BeautifulSoup

import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent))
import semanas_evisita as sem  # conversão HTML ID <-> Semana do Ano (ver módulo)

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ============================================================
# LOGGING
# ============================================================

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("evisita")


# ============================================================
# ⚙️  PARÂMETROS EDITÁVEIS — ALTERE AQUI CONFORME A NECESSIDADE
# ============================================================
# Este é o ÚNICO bloco que você normalmente precisa mexer para ajustar o
# comportamento do script. Está organizado por assunto; cada grupo tem uma
# explicação rápida do que o parâmetro faz.

# --- 1) CREDENCIAIS DE LOGIN -----------------------------------------------
# Pode ser sobrescrito por variável de ambiente (EVISITA_CPF / EVISITA_SENHA)
# ou editado direto aqui embaixo (valor após a vírgula é o padrão).
CPF   = os.environ.get("EVISITA_CPF")
SENHA = os.environ.get("EVISITA_SENHA")

if not CPF or not SENHA:
    raise SystemExit(
        "❌ Defina as variáveis de ambiente EVISITA_CPF e EVISITA_SENHA antes de rodar "
        "(no GitHub Actions isso vem dos Secrets do repositório; localmente, exporte-as "
        "no terminal antes de chamar o script)."
    )

# --- 2) FILTRO DE COLETA (o que buscar no e-Visita) ------------------------
ID_MUNICIPIO   = 62       # 62 = Ponta Porã
ID_ATIVIDADE   = 4        # 4 = Tratamento
ID_ANO         = 14       # código interno do ano epidemiológico (não é o ano civil)
ID_CICLO       = 163      # código interno do ciclo (ex.: "Rotina - 4")
ID_TIPO_VISITA = 0        # 0 = ambos (normal + recuperação) | 1 = só normal | 2 = só recuperação

# --- Cálculo automático de semana -------------------------------------------
# A conversão HTML ID <-> Semana do Ano mora inteira em scripts/semanas_evisita.py
# (é usada tanto aqui quanto no gerador do dashboard). Aqui só reaproveitamos.
ANCORA_DATA = sem.ANCORA_DATA
ANCORA_ID   = sem.ANCORA_HTML_ID
MESES_PT    = sem.MESES_PT

semana_atual        = sem.semana_atual_html    # ID de semana do site que contém hoje
_data_inicio_semana  = sem.data_inicio_semana_html
nome_periodo         = sem.nome_periodo         # rótulo amigável (nome do mês) pro seletor de período

# Primeira semana buscada na coleta inicial (quando data/semanas/ está vazio)
# — HTML 305 = Semana 18/2026 (03/05/2026), início de Maio. É INDEPENDENTE
# de ANCORA_ID/ANCORA_DATA acima (que servem só pra converter data <-> HTML
# ID em qualquer ano, não pra decidir de onde a coleta começa).
PRIMEIRA_SEMANA_COLETA = 305

# --- Coleta semana a semana --------------------------------------------------
# A partir de agora a coleta NÃO busca mais um intervalo de semanas de uma vez
# só. Cada semana (HTML ID) é buscada e cacheada SEPARADAMENTE em
# data/semanas/<html_id>.csv (+ _pe.csv pro Ponto Estratégico), e depois TODAS
# as semanas já cacheadas são consolidadas num único Analise_Consolidada.xlsx.
# Isso deixa cada execução mais rápida (só busca o que realmente mudou) e
# permite reprocessar uma semana específica sem tocar nas outras.
PASTA_SEMANAS = "data/semanas"


def semanas_ja_coletadas(pasta_semanas=PASTA_SEMANAS):
    """HTML IDs de semanas que já têm cache salvo em disco, em ordem
    crescente. Uma semana só conta como coletada se tiver o .csv principal
    (o de Ponto Estratégico é opcional — nem toda semana tem PE cadastrado)."""
    if not os.path.isdir(pasta_semanas):
        return []
    ids = []
    for f in os.listdir(pasta_semanas):
        m = re.fullmatch(r"(\d+)\.xlsx", f)
        if m:
            ids.append(int(m.group(1)))
    return sorted(ids)


def decidir_semanas_a_coletar(hoje=None, pasta_semanas=PASTA_SEMANAS):
    """Decide quais semanas (HTML IDs) buscar nesta execução:

      - Primeira execução (nada cacheado ainda): busca da semana
        PRIMEIRA_SEMANA_COLETA (Maio/2026) até a semana atual — uma
        planilha por semana.
      - Execuções seguintes: sempre RECOLETA a última semana já salva (os
        dados do e-Visita ainda podem mudar) e, se já existir uma semana
        nova disponível no site, busca ela(s) também — sem pular nenhuma,
        mesmo que o workflow tenha ficado parado por mais de uma semana.

    Retorna a lista de HTML IDs a coletar, em ordem crescente."""
    atual = semana_atual(hoje)
    ja = semanas_ja_coletadas(pasta_semanas)
    if not ja:
        log.info("📦 Primeira execução detectada (nenhuma semana em %s) — "
                  "coletando da semana %s (HTML %s) até a atual (HTML %s).",
                  pasta_semanas, sem.html_para_semana(PRIMEIRA_SEMANA_COLETA), PRIMEIRA_SEMANA_COLETA, atual)
        return list(range(PRIMEIRA_SEMANA_COLETA, atual + 1))
    ultima = max(ja)
    fim = max(atual, ultima)
    return list(range(ultima, fim + 1))

# Valores de pendência do e-Visita. O relatório principal (tempos de visita,
# alertas, etc.) é sempre baseado nos ABERTOS (ID_PENDENCIA_ABERTOS) — são os
# imóveis com visita registrada e horário real, usados em toda a análise de
# tempo/qualidade. Fechados e Recusados só entram na contagem/percentual de
# pendência (ver seção 14).
ID_PENDENCIA_ABERTOS   = 1   # imóveis com pendência em aberto (dataset principal da análise)
ID_PENDENCIA_FECHADOS  = 2   # imóveis com o ciclo fechado/concluído
ID_PENDENCIA_RECUSADOS = 3   # imóveis onde o morador recusou a visita

# --- 3) EQUIPES E AGENTES (IDs do e-Visita, agrupados por equipe) ----------
AGENTES_POR_EQUIPE = {
    "Equipe 1":        [116, 73, 1935, 470, 672, 3037, 79, 85, 30, 3101],
    "Equipe 2":        [115, 125, 3035, 3127, 2030, 134, 99, 142, 465, 81, 151, 2917],
    "Equipe 3":        [113, 2951, 120, 1938, 2721, 3128, 3038, 2918, 1936],
    "Equipe 4":        [80, 124, 161, 1943, 2789, 1996, 157, 110, 149, 86, 1942],
    "Equipe Bloqueio": [162, 1088, 1934, 109, 155, 126, 476, 122, 2101],
}

# --- 3b) PONTO ESTRATÉGICO (equipe/URL separada, com id_atividade e
# ciclo/semana próprios — ex.: id_atividade=3, tem_foto=, ciclo/semana
# geralmente diferentes do restante) --------------------------------------
# Preencha aqui os IDs dos agentes de Ponto Estratégico (mesmo formato de
# AGENTES_POR_EQUIPE, mas sem separação por equipe — se precisar separar por
# equipe no futuro, é só adaptar para um dict como o de cima).
AGENTES_PONTO_ESTRATEGICO = [
    1031,
]
ID_ATIVIDADE_PONTO_ESTRATEGICO = 3   # 3 = Ponto Estratégico (vs 4 = Tratamento no restante)

# A Atividade 3 (Ponto Estratégico) tem ciclos MENORES e independentes da
# Atividade 4 — cada ciclo novo no site tem seu próprio ID. Cadastre aqui
# cada ciclo conhecido como (semana_inicio, semana_fim, id_ciclo):
#   - ciclo 1: semanas 313-317 (28/06 a 01/08/2026) -> id_ciclo 154
#   - ciclo 2: semanas 318-321 (02/08 a 29/08/2026) -> id_ciclo 155
# Quando um novo ciclo começar no site (semana 322 em diante), adicione uma
# nova linha aqui com o id_ciclo daquele ciclo. Enquanto uma semana cair fora
# de todas as faixas cadastradas (ou o id_ciclo estiver None), a coleta do
# Ponto Estratégico é pulada com um aviso claro em vez de mandar um id_ciclo
# errado pro site.
CICLOS_PONTO_ESTRATEGICO = [
    (313, 317, 154),
    (318, 321, 155),
]

def ciclo_para_semana(semana, ciclos=CICLOS_PONTO_ESTRATEGICO):
    """Acha o id_ciclo pra uma semana específica, numa tabela de ciclos
    (serve tanto pra CICLOS_PONTO_ESTRATEGICO quanto CICLOS_TRATAMENTO).
    Retorna None se a semana cair fora de todos os ciclos cadastrados
    (ciclo futuro ainda não adicionado à lista) ou se o ciclo encontrado
    ainda não tiver id_ciclo preenchido."""
    for ini, fim, id_ciclo in ciclos:
        if ini <= semana <= fim:
            return id_ciclo
    return None

# O id_ciclo da Atividade 4 (Tratamento) também muda ao longo do tempo — não
# é fixo como o ID_CICLO abaixo fazia parecer. Confirmado com a URL real do
# site: id_ciclo=162&semana_inicio=305&semana_fim=312 (semanas 305-312,
# maio-junho/2026) — o ciclo muda a partir da semana 313, onde passa a valer
# o ID_CICLO "padrão" (163, definido mais abaixo). Faixas SEM entrada aqui
# caem no id_ciclo padrão automaticamente — só precisa cadastrar aqui os
# ciclos ANTIGOS/diferentes do padrão atual.
CICLOS_TRATAMENTO = [
    (305, 312, 162),
]

def ciclo_tratamento_para_semana(html_id, id_ciclo_padrao):
    """Como ciclo_para_semana(), mas com fallback pro id_ciclo padrão em vez
    de None — toda semana tem uma coleta de Tratamento a fazer, então não dá
    pra simplesmente pular como acontece no Ponto Estratégico."""
    for ini, fim, id_ciclo in CICLOS_TRATAMENTO:
        if ini <= html_id <= fim:
            return id_ciclo
    return id_ciclo_padrao

# (o antigo "semana_inicio_ciclo_pe" não é mais necessário: como agora cada
# semana é coletada individualmente, não existe mais a ideia de "buscar o
# ciclo inteiro desde o início numa tacada só" — cada semana busca só ela
# mesma, e o id_ciclo certo é resolvido por ciclo_para_semana() na hora.)

# --- 4) PERFORMANCE DA COLETA -----------------------------------------------
MAX_WORKERS          = 3     # quantos agentes coletar em paralelo ao mesmo tempo
DELAY_ENTRE_PAGINAS   = 0.30  # segundos de pausa entre páginas de um mesmo agente (evita sobrecarregar o servidor)
TIMEOUT_HTTP          = 60    # segundos de espera máxima por resposta do servidor antes de tentar de novo
TENTATIVAS_POR_PAGINA = 3     # tentativas extras (nível aplicação) se uma página falhar, antes de desistir dela

# --- 5) REGRAS DE ALERTA DE TEMPO DE VISITA ---------------------------------
TEMPO_MIN_VISITA_MIN = 3     # visita com duração menor que isso (min) = "rápida" (alerta)
TEMPO_MAX_VISITA_MIN = 30    # visita com duração maior que isso (min) = "longa" (alerta)
# Alguns agentes lançam pelo computador e conseguem manipular o horário de
# saída, inflando a duração da visita registrada. Pra isso não distorcer a
# média/mediana de tempo por visita, qualquer visita acima disso entra no
# cálculo de média/mediana como se tivesse durado exatamente esse teto (não
# afeta a contagem de "longas" acima, nem a listagem de alerta — só a média).
TETO_DURACAO_MEDIA_MIN = 15
META_VISITAS_DIA     = 20    # meta de visitas por dia, usada na coluna "Dias c/ Meta"

# --- 6) CLASSIFICAÇÃO DO AGENTE (🟢🟡🔴), baseada em % de visitas rápidas --
PCT_RAPIDAS_OK_ATE       = 10   # até 10% de visitas rápidas      -> 🟢 NORMAL
PCT_RAPIDAS_ATENCAO_ATE  = 20   # de 11% a 20% de visitas rápidas -> 🟡 ATENÇÃO
                                  # acima de 20%                    -> 🔴 CRÍTICO

# --- 7) SEQUÊNCIA SUSPEITA (lançamento em lote) -----------------------------
# Visitas consecutivas do mesmo agente com chegada muito próxima no tempo.
# IMPORTANTE: o e-Visita só registra minutos (sem segundos), então na prática
# esse limiar só pega visitas cravadas no MESMO MINUTO — mantido em segundos
# como critério de negócio explícito e para caso o sistema passe a fornecer
# segundos no futuro.
LIMIAR_SEQUENCIA_SUSPEITA_SEG = 50

# --- 8) HORÁRIO DE ALMOÇO (com tolerância) ----------------------------------
# Horário efetivamente monitorado como "visita durante o almoço". Já vem com
# 15 minutos de tolerância em cada borda em relação ao intervalo cheio
# (11h-13h), para não pegar visitas que só terminaram/começaram um pouco
# antes/depois do horário oficial.
HORA_ALMOCO_INICIO_MIN = 11 * 60 + 15   # 11h15 em minutos desde a meia-noite
HORA_ALMOCO_FIM_MIN    = 12 * 60 + 45   # 12h45 em minutos desde a meia-noite

# --- 9) HORÁRIO OFICIAL DE EXPEDIENTE (dias úteis) --------------------------
# Qualquer visita fora desses intervalos — inclusive sábado/domingo — entra
# na aba "Fora do Expediente".
HORA_EXPEDIENTE_MANHA_INICIO = 7
HORA_EXPEDIENTE_MANHA_FIM    = 11
HORA_EXPEDIENTE_TARDE_INICIO = 13
HORA_EXPEDIENTE_TARDE_FIM    = 17

# --- 10) CUSTO / SALÁRIO ----------------------------------------------------
SALARIO_MINIMO = 1621.00
SALARIO_AGENTE = 3 * SALARIO_MINIMO   # agente ganha 3 salários mínimos federais
DIAS_UTEIS_MES = 22                    # usado para estimar horas úteis mensais e custo da hora útil

# --- 11) SAÍDA (onde salvar o Excel) ----------------------------------------
PASTA_SAIDA         = "output"
ARQUIVO_CONSOLIDADO = "Analise_Consolidada.xlsx"

# --- 12) URLs DO SISTEMA (só mexer se o endereço do e-Visita mudar) --------
URL_LOGIN   = "https://evisita.saude.ms.gov.br/endemias"
URL_VISITAS = "https://evisita.saude.ms.gov.br/endemias/sis_visita"

# --- 13) CRONOGRAMA / AUSÊNCIAS PROGRAMADAS ---------------------------------
# Chuva, folga, atestado, reunião/treinamento e férias — tudo num só lugar,
# editável em data/cronograma_ausencias.json (use o editor visual
# cronograma_editor.html pra gerar o conteúdo, sem precisar mexer em código).
#
# Esses dias/períodos são EXCLUÍDOS do cálculo da aba "Ausências" (não
# contam como possível falta a investigar, nem entram na pontuação do
# agente) e aparecem listados na aba "Cronograma" — em cima dos dias que o
# próprio programa já detecta automaticamente como "ninguém trabalhou"
# (fins de semana, feriados nacionais já fora do range de dias úteis).
#
# Cada entrada é um dicionário com:
#   "alvo"      -> "agente" | "equipe" | "todos"
#   "id_agente" -> obrigatório só quando alvo == "agente"
#   "equipe"    -> obrigatório só quando alvo == "equipe" (nome exato, deve
#                  bater com AGENTES_POR_EQUIPE)
#   "motivo"    -> um destes cinco, sempre: "Chuva", "Folga", "Atestado",
#                  "Reunião/Treinamento", "Férias"
#   "inicio"    -> data "dd/mm/aaaa" (obrigatório)
#   "fim"       -> data "dd/mm/aaaa" (obrigatório — pra um dia só, repete a
#                  mesma data em início e fim)
MOTIVOS_CRONOGRAMA = [
    "Chuva", "Folga", "Atestado", "Reunião/Treinamento", "Férias",
    "Levantamento de Índice", "Instalação de Ovitrampa", "Recolha de Ovitrampa",
    "Recuperação de Casas",
]


def _carregar_cronograma_ausencias(caminho="data/cronograma_ausencias.json"):
    if not os.path.exists(caminho):
        return []
    try:
        with open(caminho, encoding="utf-8") as f:
            dados = json.load(f)
        if not isinstance(dados, list):
            log.warning("⚠️ %s não é uma lista — ignorando cronograma/ausências programadas.", caminho)
            return []
        return dados
    except (json.JSONDecodeError, OSError) as e:
        log.warning("⚠️ Erro ao ler %s (%s) — ignorando cronograma/ausências programadas.", caminho, e)
        return []

CRONOGRAMA_AUSENCIAS = _carregar_cronograma_ausencias()

# --- 14) COLETA DE PENDÊNCIA (fechados/recusados) ---------------------------
# Se True, o script faz coletas ADICIONAIS (uma para fechados, uma para
# recusados) além da coleta principal (abertos), para calcular o % de
# pendência e preencher a aba "Recusados e Pendências". Isso TRIPLICA o
# número de requisições feitas ao servidor — desligue (False) se quiser uma
# coleta mais rápida e não precisar dessas colunas/aba agora.
COLETAR_FECHADOS_RECUSADOS = True

# --- 15) RANKING / PONTUAÇÃO DOS AGENTES ------------------------------------
# Todo agente começa com PONTOS_INICIAIS. Cada critério abaixo desconta ou
# soma pontos conforme a quantidade de ocorrências no período analisado.
# Ajuste os pesos livremente — coloque 0 para desligar um critério sem
# precisar remover a linha.

PONTOS_INICIAIS = 100
NOTA_MINIMA = 0      # pontuação não cai abaixo disso (None = sem piso)
NOTA_MAXIMA = 120    # pontuação não passa disso (None = sem teto; >100 permite bônus se destacar)

# Horários de referência para os critérios de pontualidade por turno
# (avaliados DIA A DIA, não pela média do período).
HORA_LIMITE_INICIO_MANHA = (8, 30)    # início da manhã depois disso = atraso
HORA_LIMITE_FIM_MANHA    = (10, 0)    # fim da manhã antes disso = saiu cedo demais
HORA_LIMITE_INICIO_TARDE = (14, 30)   # início da tarde depois disso = atraso
HORA_LIMITE_FIM_TARDE    = (16, 0)    # fim da tarde antes disso = saiu cedo demais

# --- PERDAS (pontos perdidos por ocorrência) ---
PONTOS_PERDA_INICIO_MANHA_ATRASADO = 0.3   # por dia que começou a manhã após HORA_LIMITE_INICIO_MANHA
PONTOS_PERDA_FIM_MANHA_ANTECIPADO  = 0.3   # por dia que terminou a manhã antes de HORA_LIMITE_FIM_MANHA
PONTOS_PERDA_INICIO_TARDE_ATRASADO = 0.3   # por dia que começou a tarde após HORA_LIMITE_INICIO_TARDE
PONTOS_PERDA_FIM_TARDE_ANTECIPADO  = 0.3   # por dia que terminou a tarde antes de HORA_LIMITE_FIM_TARDE
PONTOS_PERDA_TURNO_SEM_LANCAMENTO  = 1.5   # por turno sem visita registrada (que não é chuva/reunião/férias)
PONTOS_PERDA_VISITA_RAPIDA         = 0.3   # por visita < TEMPO_MIN_VISITA_MIN
PONTOS_PERDA_VISITA_LONGA          = 0     # por visita > TEMPO_MAX_VISITA_MIN (desligado por padrão)
PONTOS_PERDA_VISITA_DUPLICADA      = 0.3   # por visita duplicada
PONTOS_PERDA_SEQUENCIA_SUSPEITA    = 1     # por visita envolvida em sequência suspeita (lançamento em lote)
PONTOS_PERDA_VISITA_NEGATIVA       = 0     # por visita com duração negativa (desligado — normalmente é bug do sistema, não culpa do agente)
PONTOS_PERDA_FORA_EXPEDIENTE       = 0.5   # por visita fora do expediente oficial / fim de semana
PONTOS_PERDA_VISITA_ALMOCO         = 0.5   # por visita registrada no horário de almoço
PONTOS_PERDA_DIA_SEM_META          = 0.5   # por dia trabalhado que não bateu a meta diária de visitas

# --- GANHOS (pontos ganhos por ocorrência ou bônus único no período) ---
PONTOS_GANHO_DIA_COM_META          = 0.5   # por dia que bateu a meta diária de visitas
PONTOS_GANHO_SEM_VISITA_RAPIDA     = 5     # bônus único: ZERO visitas rápidas no período inteiro
PONTOS_GANHO_SEM_AUSENCIA          = 5     # bônus único: ZERO turnos sem lançamento no período inteiro
PONTOS_GANHO_PONTUALIDADE_PERFEITA = 3     # bônus único: ZERO violações de horário (início/fim) no período

# Faixas de classificação da pontuação final
NOTA_FAIXA_EXCELENTE = 90   # >= 90        -> 🟢 EXCELENTE
NOTA_FAIXA_BOM        = 75  # 75 a 89,9    -> 🟡 BOM
NOTA_FAIXA_ATENCAO    = 60  # 60 a 74,9    -> 🟠 ATENÇÃO
                              # abaixo de 60 -> 🔴 CRÍTICO

# ============================================================
# Fim dos parâmetros editáveis — o restante do código normalmente não
# precisa ser alterado.
# ============================================================

NOMES_DIA_SEMANA = {0: "Segunda", 1: "Terça", 2: "Quarta", 3: "Quinta",
                     4: "Sexta", 5: "Sábado", 6: "Domingo"}


def _fmt_hora_min(minutos_desde_meia_noite):
    """Formata minutos-desde-meia-noite como 'HHhMM' (ex.: 675 -> '11h15')."""
    h, m = divmod(minutos_desde_meia_noite, 60)
    return f"{h:02d}h{m:02d}"


# Preposições/conectivos que ficam em minúsculo em nomes próprios (exceto se
# forem a primeira palavra do nome).
_PREPOSICOES_NOME = {"de", "da", "do", "das", "dos", "e"}


def _formatar_nome_proprio(nome):
    """
    Formata nome em Title Case (primeira letra maiúscula de cada palavra),
    mantendo preposições comuns (de/da/do/das/dos/e) em minúsculo — exceto se
    forem a primeira palavra do nome. Ex.: 'ADRIANA DE JESUS SARAIVA' ->
    'Adriana de Jesus Saraiva'.
    """
    if not nome:
        return nome
    palavras = nome.strip().lower().split()
    if not palavras:
        return nome
    resultado = []
    for i, p in enumerate(palavras):
        if i > 0 and p in _PREPOSICOES_NOME:
            resultado.append(p)
        else:
            resultado.append(p[:1].upper() + p[1:])
    return " ".join(resultado)


def _obter_motivo_cronograma(id_agente, equipe, dia_date):
    """
    Verifica se (agente, equipe, dia) bate com alguma entrada de
    CRONOGRAMA_AUSENCIAS (chuva, folga, atestado, reunião/treinamento,
    férias — alvo agente/equipe/todos). dia_date é um objeto date. O
    intervalo início-fim é INCLUSIVO nas duas pontas. Retorna o motivo
    (string) se encontrado, senão None.
    """
    for entrada in CRONOGRAMA_AUSENCIAS:
        alvo = entrada.get("alvo", "agente")
        if alvo == "agente" and entrada.get("id_agente") != id_agente:
            continue
        if alvo == "equipe" and entrada.get("equipe") != equipe:
            continue
        # alvo == "todos": sempre bate, não precisa checar mais nada
        try:
            inicio = pd.to_datetime(entrada["inicio"], format="%d/%m/%Y").date()
            fim = pd.to_datetime(entrada["fim"], format="%d/%m/%Y").date()
        except (ValueError, KeyError):
            log.warning("CRONOGRAMA_AUSENCIAS: entrada inválida (verifique o formato dd/mm/aaaa): %s", entrada)
            continue
        if inicio <= dia_date <= fim:
            return entrada.get("motivo", "Ausência programada")
    return None


def parse_args():
    """Permite sobrescrever ano/ciclo/pastas sem editar o código a cada rodada.

    --semana (HTML ID), se passado, força a coleta/recoleta de UMA semana
    específica em vez de deixar decidir_semanas_a_coletar() escolher
    automaticamente — útil pra reprocessar manualmente uma semana pontual.
    """
    p = argparse.ArgumentParser(description="Análise de tempo de visitas E-Visita")
    p.add_argument("--semana", type=int, default=None,
                    help="HTML ID de uma única semana a coletar/recoletar manualmente "
                         "(se omitido, o script decide sozinho quais semanas buscar)")
    p.add_argument("--id-ano", type=int, default=ID_ANO)
    p.add_argument("--id-ciclo", type=int, default=ID_CICLO)
    p.add_argument("--saida", type=str, default=PASTA_SAIDA)
    p.add_argument("--pasta-semanas", type=str, default=PASTA_SEMANAS)
    return p.parse_args()


# ============================================================
# LOGIN E SESSÃO
# ============================================================

def criar_driver():
    opts = Options()
    # Headless: o runner do GitHub Actions não tem tela. Localmente também
    # funciona normalmente em modo headless, então não precisa de flag extra.
    opts.add_argument("--headless=new")
    opts.add_argument("--window-size=1920,1080")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    return webdriver.Chrome(options=opts)


def fazer_login(driver):
    log.info("🔐 Realizando login...")
    driver.get(URL_LOGIN)
    wait = WebDriverWait(driver, 20)
    wait.until(EC.presence_of_element_located((By.NAME, "cpf"))).send_keys(CPF)
    wait.until(EC.presence_of_element_located((By.NAME, "password"))).send_keys(SENHA)
    wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "button[type='submit']"))).click()
    wait.until(EC.presence_of_element_located((By.TAG_NAME, "body")))
    if "login" in driver.current_url.lower():
        raise RuntimeError("Login falhou — verifique CPF/senha.")
    log.info("✅ Login OK")


def _aplicar_retry(session):
    retry = Retry(total=4, backoff_factor=0.5, status_forcelist=[429, 500, 502, 503, 504])
    adapter = HTTPAdapter(pool_connections=MAX_WORKERS, pool_maxsize=MAX_WORKERS, max_retries=retry)
    session.mount("https://", adapter)
    session.mount("http://", adapter)
    return session


def criar_session(driver):
    s = requests.Session()
    for c in driver.get_cookies():
        s.cookies.set(c["name"], c["value"], domain=c.get("domain"), path=c.get("path", "/"))
    return _aplicar_retry(s)


def clonar_session(base):
    s = requests.Session()
    for c in base.cookies:
        s.cookies.set(c.name, c.value, domain=c.domain, path=c.path)
    s.headers.update(dict(base.headers))
    return _aplicar_retry(s)


# ============================================================
# COLETA
# ============================================================

class SessaoExpiradaError(RuntimeError):
    """Levantada quando a página retornada é a tela de login (sessão caiu)."""


def montar_url(id_agente: int, pagina: int, cfg: dict, id_pendencia: int = None,
                id_atividade: int = None, tem_foto: str = None) -> str:
    params = {
        "municipio": ID_MUNICIPIO, "id_ano": cfg["id_ano"], "id_ciclo": cfg["id_ciclo"],
        "semana_inicio": cfg["semana_inicio"], "semana_fim": cfg["semana_fim"],
        "id_atividade": id_atividade if id_atividade is not None else ID_ATIVIDADE,
        "id_tipo_visita": ID_TIPO_VISITA,
        "id_pendencia": id_pendencia if id_pendencia is not None else ID_PENDENCIA_ABERTOS,
        "id_agente": id_agente,
        "acao": "filtrar", "page": pagina,
    }
    if tem_foto is not None:
        params["tem_foto"] = tem_foto
    return requests.Request("GET", URL_VISITAS, params=params).prepare().url


def obter_total_paginas(soup):
    pag = soup.find("ul", class_="pagination")
    if not pag:
        return 1
    nums = []
    for a in pag.find_all("a", class_="page-link"):
        href = a.get("href", "")
        if "page=" in href:
            try:
                nums.append(int(href.split("page=")[-1].split("&")[0]))
            except ValueError:
                log.debug("Não foi possível extrair número de página de href=%s", href)
    return max(nums, default=1)


def _extrair_nome_agente_da_pagina(soup, id_agente):
    """
    (Fallback) Extrai o nome do agente a partir do <select> de filtro presente
    na página de listagem de visitas. Só funciona se o HTML BRUTO (sem JS)
    já vier com a <option> do agente preenchida — o que pode não ser o caso
    se o site carregar essa opção via AJAX (ver extrair_nome_agente).
    """
    select = soup.find("select", id="id_agente") or soup.find("select", attrs={"name": "id_agente"})

    if select is None:
        log.debug("_extrair_nome_agente_da_pagina: <select id/name='id_agente'> não encontrado (agente %s)", id_agente)
        return None

    opt = select.find("option", attrs={"value": str(id_agente)})
    if opt is None:
        opt = next((o for o in select.find_all("option") if "selected" in o.attrs), None)

    if opt is None:
        return None

    nome = opt.get_text(strip=True)
    if not nome or nome.lower() == "selecione":
        return None
    return nome


def _obter_nome_agente_via_api(session, id_agente):
    """
    Busca o nome do agente diretamente no endpoint 'getedit' usado pelo
    JavaScript da página (select2) para popular o campo com o valor
    pré-selecionado — ex.: .../sis_usuario/getedit/672.

    Esse endpoint é chamado via AJAX pelo navegador, então o nome do agente
    normalmente NÃO aparece no HTML bruto que o requests.get() baixa (só
    depois que o JS roda no navegador). Por isso buscamos direto na fonte.

    ATENÇÃO: o formato exato da resposta desse endpoint (JSON vs HTML) não
    pôde ser confirmado sem acesso direto ao sistema — a função tenta os
    formatos mais comuns. Se não funcionar na prática, ative logging DEBUG
    e envie a resposta bruta para ajuste fino.
    """
    url = f"{URL_LOGIN}/sis_usuario/getedit/{id_agente}"
    try:
        resp = session.get(url, timeout=TIMEOUT_HTTP)
        resp.raise_for_status()
    except requests.RequestException as e:
        log.debug("getedit falhou para sis_usuario/%s: %s", id_agente, e)
        return None

    texto_bruto = resp.text.strip()

    # 1) Tenta JSON (formato mais comum para endpoints "getedit" desse tipo de sistema)
    try:
        data = resp.json()
    except ValueError:
        data = None

    if isinstance(data, dict):
        candidato = data.get("data") if isinstance(data.get("data"), dict) else data
        for chave in ("nome", "nome_usuario", "usuario", "descricao", "label", "text"):
            valor = candidato.get(chave) if isinstance(candidato, dict) else None
            if valor:
                return str(valor).strip()
        log.debug("getedit sis_usuario/%s respondeu JSON mas sem campo de nome reconhecido: %s",
                   id_agente, list(data.keys()))

    # 2) Tenta como HTML/texto simples (resposta curta = provavelmente só o nome,
    #    ou um <input value="..."> preenchido para o formulário de edição)
    if texto_bruto and "<html" not in texto_bruto.lower():
        soup_resp = BeautifulSoup(texto_bruto, "html.parser")
        input_nome = soup_resp.find("input", attrs={"name": re.compile("nome", re.I)})
        if input_nome and input_nome.get("value"):
            return input_nome["value"].strip()

        texto_limpo = soup_resp.get_text(strip=True)
        if texto_limpo and len(texto_limpo) < 100:
            return texto_limpo

    log.debug("getedit sis_usuario/%s: não foi possível extrair nome da resposta (%d caracteres)",
               id_agente, len(texto_bruto))
    return None


def extrair_nome_agente(session, soup, id_agente):
    """
    Tenta obter o nome do agente, nessa ordem:
      1) Endpoint getedit (fonte real usada pelo JS do site — mais confiável)
      2) Parsing do <select> da página de listagem (fallback, caso o site mude)
      3) Nome genérico 'Agente_{id}' (última opção, sinaliza problema no log)
    """
    nome = _obter_nome_agente_via_api(session, id_agente)
    if nome:
        return _formatar_nome_proprio(nome)

    nome = _extrair_nome_agente_da_pagina(soup, id_agente)
    if nome:
        return _formatar_nome_proprio(nome)

    log.warning("Não foi possível obter o nome do agente %s (nem via API nem via página) — "
                 "usando nome genérico. Ative logging DEBUG para investigar.", id_agente)
    return f"Agente_{id_agente}"


def extrair_visitas(soup):
    tabela = soup.find("table", class_="table")
    if not tabela or not tabela.find("tbody"):
        return []
    visitas = []
    for tr in tabela.find("tbody").find_all("tr"):
        tds = tr.find_all("td")
        if len(tds) < 9:
            continue
        link = tr.select_one("a[href*='/detalhes']")
        id_visita = None
        if link:
            m = re.search(r"/sis_visita/(\d+)/detalhes", link.get("href", "") or "")
            if m:
                id_visita = int(m.group(1))
        visitas.append({
            "id_visita": id_visita,
            "regiao": tds[0].get_text(strip=True),
            "area": tds[1].get_text(strip=True),
            "quarteirao": tds[2].get_text(strip=True),
            "logradouro": tds[3].get_text(strip=True),
            "imovel": tds[4].get_text(strip=True),
            "atividade": tds[5].get_text(strip=True),
            "tipo_visita": tds[6].get_text(strip=True),
            "data_chegada": tds[7].get_text(strip=True),
            "data_saida": tds[8].get_text(strip=True),
        })
    return visitas


def _parece_tela_login(soup):
    return soup.find("input", {"name": "cpf"}) is not None


def baixar_soup(session, url):
    resp = session.get(url, timeout=TIMEOUT_HTTP)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "html.parser")
    if _parece_tela_login(soup):
        raise SessaoExpiradaError(f"Sessão expirada ao acessar: {url}")
    return soup


def baixar_soup_com_retry(session, url, descricao="", tentativas=TENTATIVAS_POR_PAGINA):
    """
    Camada de retry em nível de APLICAÇÃO sobre baixar_soup — além do retry
    HTTP automático (urllib3, para 429/500/502/503/504), tenta de novo aqui
    para cobrir timeouts, erros de conexão e outras falhas que o retry HTTP
    não pega. Propaga SessaoExpiradaError imediatamente (não adianta insistir
    sem sessão válida).
    """
    ultimo_erro = None
    for tentativa in range(1, tentativas + 1):
        try:
            return baixar_soup(session, url)
        except SessaoExpiradaError:
            raise
        except Exception as e:
            ultimo_erro = e
            if tentativa < tentativas:
                espera = 1.0 * tentativa  # backoff simples: 1s, 2s, 3s...
                log.warning("Tentativa %s/%s falhou (%s): %s — tentando de novo em %.1fs",
                             tentativa, tentativas, descricao, e, espera)
                time.sleep(espera)
    raise ultimo_erro


def coletar_visitas_agente(session, id_agente, cfg, id_pendencia=None, obter_nome=True,
                            id_atividade=None, tem_foto=None):
    id_pendencia = id_pendencia if id_pendencia is not None else ID_PENDENCIA_ABERTOS
    log.info("  → Agente %s (pendência=%s)", id_agente, id_pendencia)
    soup_p1 = baixar_soup_com_retry(
        session, montar_url(id_agente, 1, cfg, id_pendencia, id_atividade, tem_foto),
        descricao=f"agente {id_agente} pend={id_pendencia} pág 1")
    total_pags = obter_total_paginas(soup_p1)
    todas = extrair_visitas(soup_p1)
    paginas_falhas = []

    for pag in range(2, total_pags + 1):
        try:
            soup = baixar_soup_com_retry(
                session, montar_url(id_agente, pag, cfg, id_pendencia, id_atividade, tem_foto),
                descricao=f"agente {id_agente} pend={id_pendencia} pág {pag}")
            novas = extrair_visitas(soup)
            todas.extend(novas)
            log.info("    pág %s/%s → %s visitas", pag, total_pags, len(novas))
            time.sleep(DELAY_ENTRE_PAGINAS)
        except SessaoExpiradaError:
            raise  # propaga: não adianta continuar coletando sem sessão válida
        except Exception as e:
            log.error("❌ Página %s do agente %s (pend=%s) falhou mesmo após %s tentativas: %s — "
                        "DADOS DESSA PÁGINA FICARÃO FALTANDO", pag, id_agente, id_pendencia, TENTATIVAS_POR_PAGINA, e)
            paginas_falhas.append(pag)

    nome = extrair_nome_agente(session, soup_p1, id_agente) if obter_nome else None
    por_id = {v["id_visita"]: v for v in todas if v["id_visita"]}
    finais = list(por_id.values()) + [v for v in todas if not v["id_visita"]]

    if paginas_falhas:
        log.error("⚠️ Agente %s (pend=%s): coleta INCOMPLETA — %s de %s páginas falharam (%s)",
                    id_agente, id_pendencia, len(paginas_falhas), total_pags, paginas_falhas)
    log.info("    ✅ Agente %s (pend=%s): %s visitas", id_agente, id_pendencia, len(finais))
    return nome, finais, paginas_falhas


# ============================================================
# CÁLCULOS
# ============================================================

def calcular_tempos(visitas, nome, id_agente, equipe, semana_html_id=None):
    """`semana_html_id`: quando a coleta já é feita semana a semana (o caso
    normal agora), passe o HTML ID da semana que está sendo processada — o
    df inteiro é marcado com essa semana de forma exata, sem precisar
    adivinhar pela data de cada visita. Se omitido, a semana de cada visita
    é inferida a partir da própria data de chegada (fallback, usado só se
    algum dia essa função for chamada fora do fluxo de coleta por semana)."""
    if not visitas:
        return pd.DataFrame()

    df = pd.DataFrame(visitas)
    df["id_agente"] = id_agente
    df["nome_agente"] = nome
    df["equipe"] = equipe

    # O e-Visita só fornece precisão de minutos (sem segundos) nos horários.
    df["dt_chegada"] = pd.to_datetime(df["data_chegada"], format="%d/%m/%Y %H:%M", errors="coerce")
    df["dt_saida"] = pd.to_datetime(df["data_saida"], format="%d/%m/%Y %H:%M", errors="coerce")

    n_invalidas = int(df["dt_chegada"].isna().sum() + df["dt_saida"].isna().sum())
    if n_invalidas:
        log.warning("    ⚠️ %s (%s): %s datas inválidas/não reconhecidas nas visitas",
                     nome, id_agente, n_invalidas)

    df["duracao_min"] = ((df["dt_saida"] - df["dt_chegada"]).dt.total_seconds() / 60).round(2)

    df = df.sort_values("dt_chegada").reset_index(drop=True)
    df["dia"] = df["dt_chegada"].dt.date
    # Semana do e-Visita (NÃO é semana ISO — ver scripts/semanas_evisita.py).
    # Se a chamada já sabe a semana (fluxo normal: coleta semana a semana),
    # usa ela direto; senão infere pela data de cada visita (fallback).
    if semana_html_id is not None:
        df["semana_html"] = semana_html_id
    else:
        df["semana_html"] = df["dt_chegada"].apply(
            lambda d: sem.html_id_da_data(d.date()) if pd.notna(d) else pd.NA)
    df["semana"] = df["semana_html"].apply(
        lambda h: sem.html_para_semana(int(h), sem.ano_da_semana_html(int(h))) if pd.notna(h) else pd.NA)
    df["alerta_rapida"] = df["duracao_min"] < TEMPO_MIN_VISITA_MIN
    df["alerta_longa"] = df["duracao_min"] > TEMPO_MAX_VISITA_MIN
    # Visita acima do teto usado no cálculo de média/mediana (ver
    # TETO_DURACAO_MEDIA_MIN) — listada à parte pra auditoria, mesmo que não
    # passe de TEMPO_MAX_VISITA_MIN (que é o limite de "longa").
    df["acima_teto_media"] = df["duracao_min"] > TETO_DURACAO_MEDIA_MIN
    df["alerta_neg"] = df["duracao_min"] < 0  # duração negativa = saída antes da chegada (bug do sistema)

    # Duplicadas — guarda o motivo separado (id repetido x mesmo horário+imóvel)
    df["dup_por_id"] = df["id_visita"].notna() & df.duplicated(subset=["id_visita"], keep=False)
    df["dup_por_chegada_imovel"] = df.duplicated(subset=["dt_chegada", "imovel"], keep=False)
    df["duplicada"] = df["dup_por_id"] | df["dup_por_chegada_imovel"]

    # Sequência suspeita: diferença de chegada em relação à visita anterior
    # (já ordenado por dt_chegada). Menos de LIMIAR_SEQUENCIA_SUSPEITA_SEG
    # segundos entre uma chegada e outra sugere lançamento em lote.
    diff_seg = df["dt_chegada"].diff().dt.total_seconds()
    mask_suspeita = diff_seg.between(0, LIMIAR_SEQUENCIA_SUSPEITA_SEG).fillna(False)
    df["diff_chegada_anterior_seg"] = diff_seg
    df["par_suspeito_anterior"] = mask_suspeita  # True se ESTA visita e a anterior formam par suspeito
    df["sequencia_suspeita"] = False
    idx_atual = df.index[mask_suspeita]
    df.loc[idx_atual, "sequencia_suspeita"] = True
    df.loc[idx_atual - 1, "sequencia_suspeita"] = True

    # Agrupa visitas conectadas na mesma "cadeia" suspeita (para facilitar
    # visualizar quais visitas pertencem ao mesmo lote de lançamento).
    grupo_atual = 0
    grupos = []
    for i, (susp, par_ant) in enumerate(zip(df["sequencia_suspeita"], df["par_suspeito_anterior"])):
        if not susp:
            grupos.append(None)
            continue
        if par_ant and i > 0 and grupos[i - 1] is not None:
            grupos.append(grupos[i - 1])
        else:
            grupo_atual += 1
            grupos.append(grupo_atual)
    df["grupo_sequencia_suspeita"] = grupos

    # Turnos — separa o horário de almoço (com tolerância de 15min: 11h15-12h45),
    # quando não deveria haver visitas registradas, para não contaminar as
    # estatísticas de Manhã/Tarde.
    df["hora_chegada"] = df["dt_chegada"].dt.hour
    df["minutos_chegada"] = df["dt_chegada"].dt.hour * 60 + df["dt_chegada"].dt.minute
    df["turno"] = pd.cut(
        df["minutos_chegada"], bins=[0, HORA_ALMOCO_INICIO_MIN, HORA_ALMOCO_FIM_MIN, 24 * 60],
        labels=["Manhã", "Almoço", "Tarde"], right=False,
    )
    df["visita_horario_almoco"] = df["turno"] == "Almoço"

    # Dia da semana e horário oficial de expediente (7h-11h e 13h-17h, dias
    # úteis). Qualquer visita fora disso — inclusive sábado/domingo — é
    # marcada como "fora do expediente".
    df["dia_semana_num"] = df["dt_chegada"].dt.weekday  # 0=Segunda ... 6=Domingo
    df["fim_de_semana"] = df["dia_semana_num"] >= 5
    df["dentro_expediente"] = (~df["fim_de_semana"]) & (
        ((df["hora_chegada"] >= HORA_EXPEDIENTE_MANHA_INICIO) & (df["hora_chegada"] < HORA_EXPEDIENTE_MANHA_FIM))
        | ((df["hora_chegada"] >= HORA_EXPEDIENTE_TARDE_INICIO) & (df["hora_chegada"] < HORA_EXPEDIENTE_TARDE_FIM))
    )
    df["fora_expediente"] = df["dt_chegada"].notna() & ~df["dentro_expediente"]

    return df


def _media_horario(series_dt):
    """
    Calcula o horário médio (datetime.time) de uma série de timestamps.
    pandas não sabe tirar média direto de objetos datetime.time, então
    convertemos para segundos desde a meia-noite, tiramos a média e
    convertemos de volta.
    """
    validos = series_dt.dropna()
    if validos.empty:
        return None
    segundos = validos.dt.hour * 3600 + validos.dt.minute * 60 + validos.dt.second
    media_seg = int(round(segundos.mean()))
    h, resto = divmod(media_seg, 3600)
    mi, s = divmod(resto, 60)
    return dt_time(hour=h % 24, minute=mi, second=s)


def _maior_sequencia_consecutiva(pares_suspeitos):
    """
    Recebe a lista de pares (visita[i] vs visita[i-1]) marcados como suspeitos
    e retorna o maior número de VISITAS encadeadas nessa condição.
    Ex.: pares suspeitos em 3 posições seguidas = 4 visitas na mesma cadeia.
    """
    maior_run_pares = 0
    for valor, grupo in groupby(pares_suspeitos):
        if valor:
            maior_run_pares = max(maior_run_pares, len(list(grupo)))
    return maior_run_pares + 1 if maior_run_pares > 0 else 0


def _fmt_hhmm_tupla(t):
    """Formata tupla (hora, minuto) como 'HHhMM' (ex.: (8, 30) -> '08h30')."""
    return f"{t[0]:02d}h{t[1]:02d}"


def _avaliar_pontualidade_diaria(df):
    """
    Para cada dia trabalhado, verifica se o início/fim de cada turno
    (Manhã/Tarde) violou os horários de referência configurados
    (HORA_LIMITE_INICIO_MANHA etc.) — usado pelo sistema de pontuação.
    Ignora visitas com duração negativa (horário não confiável).
    Retorna um dict com a CONTAGEM de dias com cada tipo de violação.
    """
    resultado = {
        "inicio_manha_atrasado": 0, "fim_manha_antecipado": 0,
        "inicio_tarde_atrasado": 0, "fim_tarde_antecipado": 0,
    }
    if df.empty:
        return resultado

    validos = df[~df["alerta_neg"]]
    if validos.empty:
        return resultado

    config_turnos = [
        ("Manhã", "inicio_manha_atrasado", "fim_manha_antecipado",
         HORA_LIMITE_INICIO_MANHA, HORA_LIMITE_FIM_MANHA),
        ("Tarde", "inicio_tarde_atrasado", "fim_tarde_antecipado",
         HORA_LIMITE_INICIO_TARDE, HORA_LIMITE_FIM_TARDE),
    ]

    for turno, campo_inicio, campo_fim, limite_inicio, limite_fim in config_turnos:
        sub = validos[validos["turno"] == turno]
        if sub.empty:
            continue
        por_dia = sub.groupby("dia").agg(inicio=("dt_chegada", "min"), fim=("dt_saida", "max"))
        limite_inicio_min = limite_inicio[0] * 60 + limite_inicio[1]
        limite_fim_min = limite_fim[0] * 60 + limite_fim[1]
        inicio_min = por_dia["inicio"].dt.hour * 60 + por_dia["inicio"].dt.minute
        fim_min = por_dia["fim"].dt.hour * 60 + por_dia["fim"].dt.minute
        resultado[campo_inicio] = int((inicio_min > limite_inicio_min).sum())
        resultado[campo_fim] = int((fim_min < limite_fim_min).sum())

    return resultado


def _classificar_pontuacao(pontos):
    if pontos >= NOTA_FAIXA_EXCELENTE:
        return "🟢 EXCELENTE", "D6F5D6"
    if pontos >= NOTA_FAIXA_BOM:
        return "🟡 BOM", "FFF9B0"
    if pontos >= NOTA_FAIXA_ATENCAO:
        return "🟠 ATENÇÃO", "FFD9A0"
    return "🔴 CRÍTICO", "FFB3B3"


def calcular_pontuacao(res, qtd_ausencias):
    """
    Calcula a pontuação final de um agente a partir de PONTOS_INICIAIS,
    aplicando todos os critérios de perda/ganho configurados. Retorna
    (pontos_final, classificacao, cor, lista_detalhamento).
    """
    r = res["resumo"]
    df = res["df"]
    pontos = float(PONTOS_INICIAIS)
    detalhes = []

    def aplicar(motivo, qtd, peso, sinal):
        nonlocal pontos
        if not qtd or not peso:
            return
        delta = sinal * qtd * peso
        pontos += delta
        detalhes.append({
            "nome_agente": res["nome_agente"],
            "equipe": res["equipe"],
            "motivo": motivo,
            "qtd": qtd,
            "peso": peso,
            "pontos": round(delta, 2),
        })

    pont_diaria = _avaliar_pontualidade_diaria(df)
    aplicar(f"Início manhã após {_fmt_hhmm_tupla(HORA_LIMITE_INICIO_MANHA)}",
            pont_diaria["inicio_manha_atrasado"], PONTOS_PERDA_INICIO_MANHA_ATRASADO, -1)
    aplicar(f"Fim manhã antes de {_fmt_hhmm_tupla(HORA_LIMITE_FIM_MANHA)}",
            pont_diaria["fim_manha_antecipado"], PONTOS_PERDA_FIM_MANHA_ANTECIPADO, -1)
    aplicar(f"Início tarde após {_fmt_hhmm_tupla(HORA_LIMITE_INICIO_TARDE)}",
            pont_diaria["inicio_tarde_atrasado"], PONTOS_PERDA_INICIO_TARDE_ATRASADO, -1)
    aplicar(f"Fim tarde antes de {_fmt_hhmm_tupla(HORA_LIMITE_FIM_TARDE)}",
            pont_diaria["fim_tarde_antecipado"], PONTOS_PERDA_FIM_TARDE_ANTECIPADO, -1)

    aplicar("Turno sem lançamento (não é chuva/reunião/férias)",
            qtd_ausencias, PONTOS_PERDA_TURNO_SEM_LANCAMENTO, -1)
    aplicar("Visita rápida", r["visitas_rapidas"], PONTOS_PERDA_VISITA_RAPIDA, -1)
    aplicar("Visita longa", r["visitas_longas"], PONTOS_PERDA_VISITA_LONGA, -1)
    aplicar("Visita duplicada", r["visitas_duplicadas"], PONTOS_PERDA_VISITA_DUPLICADA, -1)
    aplicar("Visita em sequência suspeita", r["visitas_seq_suspeita"], PONTOS_PERDA_SEQUENCIA_SUSPEITA, -1)
    aplicar("Visita com duração negativa", r["visitas_negativas"], PONTOS_PERDA_VISITA_NEGATIVA, -1)
    aplicar("Visita fora do expediente/fim de semana", r["visitas_fora_expediente"], PONTOS_PERDA_FORA_EXPEDIENTE, -1)
    aplicar("Visita no horário de almoço", r["visitas_horario_almoco"], PONTOS_PERDA_VISITA_ALMOCO, -1)

    dias_sem_meta = max(r["dias_trabalhados"] - r["dias_atingiu_meta"], 0)
    aplicar("Dia sem bater a meta diária", dias_sem_meta, PONTOS_PERDA_DIA_SEM_META, -1)

    aplicar("Dia com meta diária batida", r["dias_atingiu_meta"], PONTOS_GANHO_DIA_COM_META, +1)

    if r["visitas_rapidas"] == 0 and r["total_visitas"] > 0:
        aplicar("Bônus: nenhuma visita rápida no período", 1, PONTOS_GANHO_SEM_VISITA_RAPIDA, +1)
    if qtd_ausencias == 0 and r["dias_trabalhados"] > 0:
        aplicar("Bônus: nenhum turno sem lançamento no período", 1, PONTOS_GANHO_SEM_AUSENCIA, +1)
    if sum(pont_diaria.values()) == 0 and r["dias_trabalhados"] > 0:
        aplicar("Bônus: pontualidade perfeita no período", 1, PONTOS_GANHO_PONTUALIDADE_PERFEITA, +1)

    if NOTA_MINIMA is not None:
        pontos = max(pontos, NOTA_MINIMA)
    if NOTA_MAXIMA is not None:
        pontos = min(pontos, NOTA_MAXIMA)

    pontos = round(pontos, 2)
    classif, fill = _classificar_pontuacao(pontos)
    return pontos, classif, fill, detalhes


def _horas_trabalhadas_diarias(df):
    """
    Para cada dia trabalhado, calcula:
      - horas_span: jornada total (primeira chegada até última saída do dia),
        DESCONTANDO a sobreposição com o horário de almoço (o servidor não
        deveria estar em expediente nesse intervalo).
      - horas_uteis: soma das durações de visita naquele dia (tempo
        efetivamente dentro de imóveis/atividade), em horas.
    Visitas com duração negativa são ignoradas (horário de saída não confiável).
    """
    validos = df[~df["alerta_neg"]]
    if validos.empty:
        return pd.DataFrame(columns=["horas_span", "horas_uteis"])

    agg = validos.groupby("dia").agg(
        inicio=("dt_chegada", "min"),
        fim=("dt_saida", "max"),
        soma_min=("duracao_min", "sum"),
    )

    def _overlap_almoco_horas(row):
        dia = row.name
        h_ini, m_ini = divmod(HORA_ALMOCO_INICIO_MIN, 60)
        h_fim, m_fim = divmod(HORA_ALMOCO_FIM_MIN, 60)
        almoco_ini = pd.Timestamp.combine(dia, dt_time(hour=h_ini, minute=m_ini))
        almoco_fim = pd.Timestamp.combine(dia, dt_time(hour=h_fim, minute=m_fim))
        ini_sobreposto = max(row["inicio"], almoco_ini)
        fim_sobreposto = min(row["fim"], almoco_fim)
        segundos = (fim_sobreposto - ini_sobreposto).total_seconds()
        return max(segundos, 0) / 3600

    agg["overlap_almoco_h"] = agg.apply(_overlap_almoco_horas, axis=1)
    agg["horas_span"] = ((agg["fim"] - agg["inicio"]).dt.total_seconds() / 3600) - agg["overlap_almoco_h"]
    agg["horas_span"] = agg["horas_span"].clip(lower=0)
    agg["horas_uteis"] = agg["soma_min"] / 60

    return agg[["horas_span", "horas_uteis"]]




def resumo_agente(df):
    # Média/mediana de duração EXCLUEM visitas com duração negativa (bug do
    # sistema — saída registrada antes da chegada). Ver aba "Visitas Negativas".
    dv = df[df["duracao_min"].notna() & (df["duracao_min"] >= 0)].copy()
    dv["duracao_capada"] = dv["duracao_min"].clip(upper=TETO_DURACAO_MEDIA_MIN)
    por_dia = df.groupby("dia")["imovel"].count()

    total_duplicadas = int(df["duplicada"].sum())
    pct_duplicadas = round((total_duplicadas / len(df) * 100), 2) if len(df) > 0 else 0

    total_seq_suspeita = int(df["sequencia_suspeita"].sum())
    pct_seq_suspeita = round((total_seq_suspeita / len(df) * 100), 2) if len(df) > 0 else 0
    maior_seq_suspeita = _maior_sequencia_consecutiva(df["par_suspeito_anterior"].tolist())

    total_almoco = int(df["visita_horario_almoco"].sum())
    pct_almoco = round((total_almoco / len(df) * 100), 2) if len(df) > 0 else 0

    total_fora_expediente = int(df["fora_expediente"].sum())
    pct_fora_expediente = round((total_fora_expediente / len(df) * 100), 2) if len(df) > 0 else 0

    # Horas trabalhadas (jornada) e horas úteis (tempo em visita), por dia e por semana
    diario_horas = _horas_trabalhadas_diarias(df)
    media_horas_trabalhadas_dia = round(diario_horas["horas_span"].mean(), 2) if not diario_horas.empty else 0
    media_horas_uteis_dia_exata = diario_horas["horas_uteis"].mean() if not diario_horas.empty else 0
    media_horas_uteis_dia = round(media_horas_uteis_dia_exata, 2) if not diario_horas.empty else 0

    if not diario_horas.empty:
        dia_para_semana = df.groupby("dia")["semana"].first()
        horas_por_semana = diario_horas["horas_span"].groupby(dia_para_semana).sum()
        media_horas_trabalhadas_semana = round(horas_por_semana.mean(), 2) if not horas_por_semana.empty else 0
    else:
        media_horas_trabalhadas_semana = 0

    # Custo da hora útil: salário mensal fixo / horas úteis mensais estimadas
    # (média diária de horas em visita, SEM arredondar, extrapolada para
    # DIAS_UTEIS_MES dias — arredondar antes distorceria o custo final).
    horas_uteis_mes_estimado = media_horas_uteis_dia_exata * DIAS_UTEIS_MES
    custo_hora_util = round(SALARIO_AGENTE / horas_uteis_mes_estimado, 2) if horas_uteis_mes_estimado > 0 else None

    # Horários por turno: início = primeira chegada do dia; fim = última saída
    # do dia; depois tira-se a média dessas primeiras/últimas entre os dias
    # trabalhados (em vez de fazer a média de TODOS os horários de chegada,
    # o que incluía visitas do meio do turno e distorcia o resultado).
    stats = {}
    for t in ["Manhã", "Tarde"]:
        sub = df[(df["turno"] == t) & (~df["alerta_neg"])]
        if not sub.empty:
            por_dia_turno = sub.groupby("dia").agg(inicio=("dt_chegada", "min"), fim=("dt_saida", "max"))
            stats[f"inicio_{t.lower()}"] = _media_horario(por_dia_turno["inicio"])
            stats[f"fim_{t.lower()}"] = _media_horario(por_dia_turno["fim"])
        else:
            stats[f"inicio_{t.lower()}"] = None
            stats[f"fim_{t.lower()}"] = None

    return {
        "total_visitas": len(df),
        "dias_trabalhados": df["dia"].nunique(),
        "media_min": round(dv["duracao_capada"].mean(), 2) if not dv.empty else 0,
        "mediana_min": round(dv["duracao_capada"].median(), 2) if not dv.empty else 0,
        "visitas_rapidas": int(df["alerta_rapida"].sum()),
        "visitas_longas": int(df["alerta_longa"].sum()),
        "visitas_negativas": int(df["alerta_neg"].sum()),
        "pct_rapidas": round(df["alerta_rapida"].mean() * 100, 2),
        "visitas_por_dia": round(len(df) / max(df["dia"].nunique(), 1), 2),
        "dias_atingiu_meta": int((por_dia >= META_VISITAS_DIA).sum()),
        "visitas_duplicadas": total_duplicadas,
        "pct_duplicadas": pct_duplicadas,
        "visitas_seq_suspeita": total_seq_suspeita,
        "pct_seq_suspeita": pct_seq_suspeita,
        "maior_seq_suspeita": maior_seq_suspeita,
        "visitas_horario_almoco": total_almoco,
        "pct_horario_almoco": pct_almoco,
        "visitas_fora_expediente": total_fora_expediente,
        "pct_fora_expediente": pct_fora_expediente,
        "media_horas_trabalhadas_dia": media_horas_trabalhadas_dia,
        "media_horas_trabalhadas_semana": media_horas_trabalhadas_semana,
        "media_horas_uteis_dia": media_horas_uteis_dia,
        "custo_hora_util": custo_hora_util,
        **stats,
    }


def resumo_por_area(df, fechados_por_area=None, recusados_por_area=None):
    if df.empty:
        return pd.DataFrame()
    fechados_por_area = fechados_por_area or {}
    recusados_por_area = recusados_por_area or {}

    # Exclui visitas com duração negativa do cálculo de tempo médio/mediana
    # (mesmo motivo do resumo por agente — bug do sistema, ver aba "Visitas Negativas").
    dv = df[df["duracao_min"].notna() & (df["duracao_min"] >= 0)].copy()
    dv["duracao_capada"] = dv["duracao_min"].clip(upper=TETO_DURACAO_MEDIA_MIN)
    agg = dv.groupby("area").agg(
        qtd_visitas=("imovel", "count"),
        tempo_medio=("duracao_capada", "mean"),
        tempo_mediana=("duracao_capada", "median"),
        rapidas=("alerta_rapida", "sum")
    ).reset_index()

    q_lista = df.groupby("area")["quarteirao"].apply(
        lambda x: ", ".join(sorted(set(x.dropna().astype(str))))
    ).reset_index(name="quarteiroes")
    q_qtd = df.groupby("area")["quarteirao"].nunique().reset_index(name="qtd_quarteiroes")

    agg = agg.merge(q_lista, on="area").merge(q_qtd, on="area")
    agg["tempo_medio"] = agg["tempo_medio"].round(2)
    agg["tempo_mediana"] = agg["tempo_mediana"].round(2)
    agg["pct_rapidas"] = (agg["rapidas"] / agg["qtd_visitas"] * 100).round(2)

    agg["qtd_fechados"] = agg["area"].map(fechados_por_area).fillna(0).astype(int)
    agg["qtd_recusados"] = agg["area"].map(recusados_por_area).fillna(0).astype(int)
    total_geral_area = agg["qtd_visitas"] + agg["qtd_fechados"] + agg["qtd_recusados"]
    agg["pct_pendencia_area"] = ((agg["qtd_fechados"] + agg["qtd_recusados"]) / total_geral_area.replace(0, pd.NA) * 100).round(2).fillna(0)

    agg["alerta"] = agg["pct_rapidas"].apply(lambda p: _classificar(p)[0])
    return agg.sort_values("tempo_medio")


# ============================================================
# PLANILHA
# ============================================================

COLS_RESUMO = [
    ("Agente", "nome_agente"),
    ("Equipe", "equipe"),
    ("Classificação", "classificacao"),
    ("Imóveis Abertos", "total_visitas"),
    ("Dias Trabalhados", "dias_trabalhados"),
    ("Visitas/Dia", "visitas_por_dia"),
    ("Média (min)", "media_min"),
    ("Mediana (min)", "mediana_min"),
    ("Rápidas", "visitas_rapidas"),
    ("% Rápidas", "pct_rapidas"),
    ("Longas", "visitas_longas"),
    ("Negativas", "visitas_negativas"),
    ("Dias c/ Meta", "dias_atingiu_meta"),
    ("Visitas Duplicadas", "visitas_duplicadas"),
    ("Visitas em Sequência Suspeita", "visitas_seq_suspeita"),
    ("Visitas no Almoço (11h15-12h45)", "visitas_horario_almoco"),
    ("Fora do Expediente/Fim de Semana", "visitas_fora_expediente"),
    ("Imóveis Fechados", "imoveis_fechados"),
    ("Imóveis Recusados", "imoveis_recusados"),
    ("Total Geral (Abertos+Fechados+Recusados)", "total_geral_pendencia"),
    ("% Pendência", "pct_pendencia"),
    ("Média Horas Trabalhadas/Dia", "media_horas_trabalhadas_dia"),
    ("Média Horas Trabalhadas/Semana", "media_horas_trabalhadas_semana"),
    ("Média Horas Úteis/Dia", "media_horas_uteis_dia"),
    ("Custo Hora Útil (R$)", "custo_hora_util"),
    ("Média Início Manhã", "inicio_manhã"),
    ("Média Fim Manhã", "fim_manhã"),
    ("Média Início Tarde", "inicio_tarde"),
    ("Média Fim Tarde", "fim_tarde"),
]

COLS_RANKING = [
    ("Posição", "posicao"),
    ("Agente", "nome_agente"),
    ("Equipe", "equipe"),
    ("Pontuação Final", "pontos"),
    ("Classificação", "classificacao"),
]

COLS_RANKING_DETALHE = [
    ("Agente", "nome_agente"),
    ("Equipe", "equipe"),
    ("Critério", "motivo"),
    ("Quantidade", "qtd"),
    ("Peso (pontos/ocorrência)", "peso"),
    ("Pontos Aplicados", "pontos"),
]

COLS_PONTO_ESTRATEGICO = [
    ("Semana", "semana"),
    ("Período", "periodo_semana"),
] + COLS_RESUMO

COLS_AREA = [
    ("Agente", "nome_agente"),
    ("Equipe", "equipe"),
    ("Área", "area"),
    ("Qtd Imóveis Abertos", "qtd_visitas"),
    ("Tempo Médio (min)", "tempo_medio"),
    ("Mediana (min)", "tempo_mediana"),
    ("Qtd Quarteirões", "qtd_quarteiroes"),
    ("Quarteirões", "quarteiroes"),
    ("Rápidas", "rapidas"),
    ("% Rápidas", "pct_rapidas"),
    ("Qtd Fechados", "qtd_fechados"),
    ("Qtd Recusados", "qtd_recusados"),
    ("% Pendência", "pct_pendencia_area"),
    ("Alerta", "alerta"),
]

COLS_DUPLICADAS = [
    ("Agente", "nome_agente"),
    ("Equipe", "equipe"),
    ("ID Visita", "id_visita"),
    ("Região", "regiao"),
    ("Área", "area"),
    ("Quarteirão", "quarteirao"),
    ("Logradouro", "logradouro"),
    ("Imóvel", "imovel"),
    ("Data Chegada", "data_chegada"),
    ("Data Saída", "data_saida"),
    ("Motivo", "motivo_duplicacao"),
]

COLS_SUSPEITAS = [
    ("Agente", "nome_agente"),
    ("Equipe", "equipe"),
    ("Cadeia", "grupo_sequencia_suspeita"),
    ("ID Visita", "id_visita"),
    ("Região", "regiao"),
    ("Área", "area"),
    ("Quarteirão", "quarteirao"),
    ("Logradouro", "logradouro"),
    ("Imóvel", "imovel"),
    ("Data Chegada", "data_chegada"),
    ("Data Saída", "data_saida"),
    ("Diferença p/ Anterior (s)*", "diff_chegada_anterior_seg"),
]

COLS_NEGATIVAS = [
    ("Agente", "nome_agente"),
    ("Equipe", "equipe"),
    ("ID Visita", "id_visita"),
    ("Região", "regiao"),
    ("Área", "area"),
    ("Quarteirão", "quarteirao"),
    ("Logradouro", "logradouro"),
    ("Imóvel", "imovel"),
    ("Data Chegada", "data_chegada"),
    ("Data Saída", "data_saida"),
    ("Duração (min)", "duracao_min"),
    ("Observação", "observacao"),
]

COLS_ACIMA_TETO_MEDIA = [
    ("Agente", "nome_agente"),
    ("Equipe", "equipe"),
    ("ID Visita", "id_visita"),
    ("Região", "regiao"),
    ("Área", "area"),
    ("Quarteirão", "quarteirao"),
    ("Logradouro", "logradouro"),
    ("Imóvel", "imovel"),
    ("Data Chegada", "data_chegada"),
    ("Data Saída", "data_saida"),
    ("Duração (min)", "duracao_min"),
]

COLS_FORA_EXPEDIENTE = [
    ("Agente", "nome_agente"),
    ("Equipe", "equipe"),
    ("ID Visita", "id_visita"),
    ("Dia da Semana", "dia_semana"),
    ("Região", "regiao"),
    ("Área", "area"),
    ("Quarteirão", "quarteirao"),
    ("Logradouro", "logradouro"),
    ("Imóvel", "imovel"),
    ("Data Chegada", "data_chegada"),
    ("Data Saída", "data_saida"),
    ("Motivo", "motivo"),
]

COLS_ALMOCO = [
    ("Agente", "nome_agente"),
    ("Equipe", "equipe"),
    ("ID Visita", "id_visita"),
    ("Região", "regiao"),
    ("Área", "area"),
    ("Quarteirão", "quarteirao"),
    ("Logradouro", "logradouro"),
    ("Imóvel", "imovel"),
    ("Data Chegada", "data_chegada"),
    ("Data Saída", "data_saida"),
]

COLS_AUSENCIAS = [
    ("Agente", "nome_agente"),
    ("Equipe", "equipe"),
    ("Dia", "dia_fmt"),
    ("Dia da Semana", "dia_semana"),
    ("Turno", "turno"),
    ("Observação", "observacao"),
]

COLS_CRONOGRAMA = [
    ("Dia Início", "dia_inicio"),
    ("Dia Fim", "dia_fim"),
    ("Motivo", "motivo"),
    ("Quem", "alvo_fmt"),
]

COLS_PENDENCIAS_RESUMO = [
    ("Agente", "nome_agente"),
    ("Equipe", "equipe"),
    ("Imóveis Abertos", "abertos"),
    ("Imóveis Fechados", "fechados"),
    ("Imóveis Recusados", "recusados"),
    ("Total Geral", "total_geral"),
    ("% Pendência", "pct_pendencia"),
]

COLS_RECUSADOS_DETALHE = [
    ("Agente", "nome_agente"),
    ("Equipe", "equipe"),
    ("ID Visita", "id_visita"),
    ("Região", "regiao"),
    ("Área", "area"),
    ("Quarteirão", "quarteirao"),
    ("Logradouro", "logradouro"),
    ("Imóvel", "imovel"),
    ("Data Chegada", "data_chegada"),
    ("Data Saída", "data_saida"),
]


def _classificar(pct_rapidas):
    """
    Classificação baseada no % DE VISITAS RÁPIDAS (< TEMPO_MIN_VISITA_MIN):
      - até 10%      -> 🟢 NORMAL/OK
      - 11% a 20%    -> 🟡 ATENÇÃO
      - acima de 20% -> 🔴 CRÍTICO
    """
    if pct_rapidas <= PCT_RAPIDAS_OK_ATE:
        return "🟢 NORMAL", "D6F5D6"
    if pct_rapidas <= PCT_RAPIDAS_ATENCAO_ATE:
        return "🟡 ATENÇÃO", "FFF9B0"
    return "🔴 CRÍTICO", "FFB3B3"


def _estilizar(ws, linha_cab: int, cor_cab="2E4057", titulo=None, ultima_linha_dados=None):
    lado = Side(style="thin")
    borda = Border(left=lado, right=lado, top=lado, bottom=lado)

    if titulo:
        n_cols = ws.max_column
        ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
        c = ws["A1"]
        c.value = titulo
        c.font = Font(bold=True, size=12, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1A1A2E")
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 26

    for cell in ws[linha_cab]:
        if cell.value is None:
            continue
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor=cor_cab)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[linha_cab].height = 28

    for row in ws.iter_rows():
        for cell in row:
            cell.border = borda

    # Autoajuste de largura: mede o maior conteúdo de cada coluna a partir da
    # linha 2 (pula a linha 1, que é o título mesclado e distorceria a coluna A).
    for col in range(1, ws.max_column + 1):
        larg = max(
            (len(str(row[0].value or "")) for row in ws.iter_rows(min_row=2, min_col=col, max_col=col)),
            default=8,
        )
        ws.column_dimensions[get_column_letter(col)].width = max(8, min(larg + 2, 45))

    # Autofiltro: da linha de cabeçalho até a última linha de dados
    if ultima_linha_dados and ultima_linha_dados >= linha_cab:
        n_cols = ws.max_column
        ws.auto_filter.ref = f"A{linha_cab}:{get_column_letter(n_cols)}{ultima_linha_dados}"


def _adicionar_legenda(ws, linha_inicio: int):
    """Adiciona bloco de legenda explicando a classificação por % de visitas rápidas e outros alertas."""
    itens = [
        ("🟢 NORMAL", "D6F5D6", f"% Rápidas até {PCT_RAPIDAS_OK_ATE}%"),
        ("🟡 ATENÇÃO", "FFF9B0", f"% Rápidas entre {PCT_RAPIDAS_OK_ATE + 1}% e {PCT_RAPIDAS_ATENCAO_ATE}%"),
        ("🔴 CRÍTICO", "FFB3B3", f"% Rápidas acima de {PCT_RAPIDAS_ATENCAO_ATE}%"),
    ]

    linha = linha_inicio
    cab = ws.cell(row=linha, column=1, value="LEGENDA — CLASSIFICAÇÃO (baseada em % de Visitas Rápidas)")
    cab.font = Font(bold=True, size=10, color="FFFFFF")
    cab.fill = PatternFill("solid", fgColor="2E4057")
    n_cols_legenda = 3
    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=n_cols_legenda)
    cab.alignment = Alignment(horizontal="left", vertical="center")
    linha += 1

    lado = Side(style="thin")
    borda = Border(left=lado, right=lado, top=lado, bottom=lado)
    for rotulo, cor, descricao in itens:
        c1 = ws.cell(row=linha, column=1, value=rotulo)
        c1.fill = PatternFill("solid", fgColor=cor)
        c1.font = Font(bold=True, size=10)
        c1.alignment = Alignment(horizontal="center", vertical="center")
        c1.border = borda

        c2 = ws.cell(row=linha, column=2, value=descricao)
        c2.alignment = Alignment(horizontal="left", vertical="center")
        c2.border = borda
        ws.merge_cells(start_row=linha, start_column=2, end_row=linha, end_column=n_cols_legenda)
        linha += 1

    linha += 1  # linha em branco

    # Explicação da Sequência Suspeita
    cab2 = ws.cell(row=linha, column=1, value="⚠️ SEQUÊNCIA SUSPEITA")
    cab2.font = Font(bold=True, size=10, color="FFFFFF")
    cab2.fill = PatternFill("solid", fgColor="7A4B00")
    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=n_cols_legenda)
    cab2.alignment = Alignment(horizontal="left", vertical="center")
    linha += 1

    desc = ws.cell(
        row=linha, column=1,
        value=("Duas ou mais visitas consecutivas do mesmo agente registradas no MESMO "
                "MINUTO de chegada são marcadas como suspeitas (o e-Visita não registra "
                "segundos, então o critério de negócio de "
                f"{LIMIAR_SEQUENCIA_SUSPEITA_SEG}s de diferença equivale, na prática, a "
                "'mesmo minuto') — sugere lançamento em lote no sistema, não visita real "
                "de campo. Colunas: 'Visitas em Sequência Suspeita' (total de visitas "
                "envolvidas) e 'Maior Sequência Suspeita' (maior número de visitas "
                "seguidas nessa condição). Na aba 'Visitas Suspeitas', a coluna "
                "'Diferença p/ Anterior (s)*' fica em branco quando o alerta da "
                "linha veio do par com a visita SEGUINTE (não da anterior).")
    )
    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=n_cols_legenda)
    desc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[linha].height = 45
    linha += 1

    linha += 1  # linha em branco

    # Explicação do Horário de Almoço
    cab3 = ws.cell(row=linha, column=1, value="🍽️ VISITAS NO HORÁRIO DE ALMOÇO")
    cab3.font = Font(bold=True, size=10, color="FFFFFF")
    cab3.fill = PatternFill("solid", fgColor="1F6F5C")
    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=n_cols_legenda)
    cab3.alignment = Alignment(horizontal="left", vertical="center")
    linha += 1

    desc3 = ws.cell(
        row=linha, column=1,
        value=(f"Coluna 'Visitas no Almoço ({_fmt_hora_min(HORA_ALMOCO_INICIO_MIN)}-"
                f"{_fmt_hora_min(HORA_ALMOCO_FIM_MIN)})' conta as visitas com chegada registrada "
                f"nesse intervalo — o horário oficial de almoço é {_fmt_hora_min(HORA_ALMOCO_INICIO_MIN)} "
                f"às {_fmt_hora_min(HORA_ALMOCO_FIM_MIN)}, JÁ COM 15 MINUTOS DE TOLERÂNCIA em cada "
                "borda (visitas entre 11h-11h15 ou 12h45-13h não são tratadas como violação, só como "
                "'fora do expediente' genérico) — use para checar se o agente está trabalhando no "
                "horário de almoço ou se há erro de lançamento. Essas visitas ficam de fora do cálculo "
                "de 'Início/Fim Manhã' e 'Início/Fim Tarde', que usam a primeira chegada e a última "
                "saída de cada dia (não a média de todos os horários), para refletir melhor o início "
                "e o fim reais do expediente.")
    )
    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=n_cols_legenda)
    desc3.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[linha].height = 70
    linha += 1

    linha += 1  # linha em branco

    # Explicação de Horas Trabalhadas / Horas Úteis / Custo da Hora Útil
    cab4 = ws.cell(row=linha, column=1, value="💰 HORAS TRABALHADAS, HORAS ÚTEIS E CUSTO DA HORA ÚTIL")
    cab4.font = Font(bold=True, size=10, color="FFFFFF")
    cab4.fill = PatternFill("solid", fgColor="4A6741")
    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=n_cols_legenda)
    cab4.alignment = Alignment(horizontal="left", vertical="center")
    linha += 1

    desc4 = ws.cell(
        row=linha, column=1,
        value=("'Horas Trabalhadas' = jornada total do dia (primeira chegada até a última "
                "saída), descontando a sobreposição com o horário de almoço — depois tira-se "
                "a média entre os dias/semanas trabalhados. 'Horas Úteis' = soma das durações "
                "de visita no dia (tempo realmente dentro de imóveis), sem contar deslocamento "
                "entre visitas. 'Custo Hora Útil' = salário mensal do agente "
                f"(R$ {SALARIO_AGENTE:.2f}".replace(".", ",") + f", equivalente a 3 salários "
                f"mínimos federais) dividido pelas horas úteis mensais estimadas (Média Horas "
                f"Úteis/Dia × {DIAS_UTEIS_MES} dias úteis/mês). Quanto menos tempo o agente "
                "passa efetivamente em visita por dia, mais caro fica o custo por hora útil.")
    )
    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=n_cols_legenda)
    desc4.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[linha].height = 60
    linha += 1

    linha += 1  # linha em branco

    # Nota sobre a aba de duplicadas
    nota = ws.cell(
        row=linha, column=1,
        value=("ℹ️ A lista detalhada de visitas duplicadas (com endereço) está na aba "
               "'Visitas Duplicadas' — use-a para localizar e excluir os registros repetidos.")
    )
    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=n_cols_legenda)
    nota.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    nota.font = Font(italic=True, size=9)
    linha += 1

    # Nota sobre a aba de negativas
    nota2 = ws.cell(
        row=linha, column=1,
        value=("ℹ️ Visitas com duração NEGATIVA (saída antes da chegada — provável bug do "
               "sistema) foram excluídas do cálculo de Média/Mediana e estão detalhadas na "
               "aba 'Visitas Negativas', para investigação junto ao suporte do e-Visita.")
    )
    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=n_cols_legenda)
    nota2.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    nota2.font = Font(italic=True, size=9)
    linha += 1

    return linha


def _escrever(ws, linha, registros, colunas):
    for dado in registros:
        fill = dado.get("_fill", "")
        for col, (_, chave) in enumerate(colunas, 1):
            valor = dado.get(chave, "")
            if isinstance(valor, bool):
                valor = "Sim" if valor else "Não"
            if chave == "id_visita" and valor not in ("", None):
                valor = str(int(valor)) if isinstance(valor, float) else str(valor)
            cell = ws.cell(row=linha, column=col, value=valor)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if chave == "id_visita":
                cell.number_format = "@"  # texto — evita notação científica e erros de conversão
                if valor not in ("", None) and str(valor).isdigit():
                    cell.hyperlink = f"{URL_VISITAS}/{valor}/detalhes"
                    cell.font = Font(color="0563C1", underline="single")
            if fill:
                cell.fill = PatternFill("solid", fgColor=fill)
        linha += 1
    return linha


def _formatar_campos_resumo(resumo_dict):
    """
    Formata os campos calculados de um dicionário 'resumo' (vindo de
    resumo_agente): classificação por % rápidas, horários HH:MM, custo em R$.
    Reaproveitado tanto pelo Resumo Geral quanto pela aba Ponto Estratégico
    (que aplica o mesmo cálculo por semana).
    """
    r = dict(resumo_dict)  # cópia — evita mutar o dict original em memória
    classif, fill = _classificar(r.get("pct_rapidas", 0))

    for key in ["inicio_manhã", "fim_manhã", "inicio_tarde", "fim_tarde"]:
        if r.get(key):
            try:
                r[key] = r[key].strftime("%H:%M")
            except AttributeError:
                r[key] = str(r[key])

    if r.get("custo_hora_util") is not None:
        r["custo_hora_util"] = f"R$ {r['custo_hora_util']:.2f}".replace(".", ",")
    else:
        r["custo_hora_util"] = "—"

    return r, classif, fill


def _preparar_resumo(res):
    r, classif, fill = _formatar_campos_resumo(res["resumo"])
    return {
        **r,
        "nome_agente": res["nome_agente"],
        "equipe": res["equipe"],
        "classificacao": classif,
        "_fill": fill,
    }


def _preparar_area(df_ag, res):
    df_area = resumo_por_area(df_ag, res.get("fechados_por_area"), res.get("recusados_por_area"))
    if df_area.empty:
        return []
    registros = []
    for _, row in df_area.iterrows():
        _, fill = _classificar(row["pct_rapidas"])
        registros.append({
            **row.to_dict(),
            "nome_agente": res["nome_agente"],
            "equipe": res["equipe"],
            "_fill": fill,
        })
    return registros


def _preparar_duplicadas(df_total):
    """
    Monta a lista de visitas duplicadas com endereço completo, para facilitar
    localizar e excluir os registros repetidos diretamente no sistema.
    """
    if df_total.empty:
        return []

    dup = df_total[df_total["duplicada"]].copy()
    if dup.empty:
        return []

    def _motivo(row):
        motivos = []
        if row["dup_por_id"]:
            motivos.append("ID de visita repetido")
        if row["dup_por_chegada_imovel"]:
            motivos.append("Mesmo horário de chegada + imóvel")
        return " / ".join(motivos) if motivos else "Duplicada"

    dup["motivo_duplicacao"] = dup.apply(_motivo, axis=1)

    # Agrupa visualmente as duplicatas: mesmo agente + imóvel + chegada ficam juntas
    dup = dup.sort_values(["nome_agente", "area", "imovel", "dt_chegada"])

    registros = []
    for _, row in dup.iterrows():
        registros.append({
            "nome_agente": row["nome_agente"],
            "equipe": row["equipe"],
            "id_visita": row["id_visita"] if pd.notna(row["id_visita"]) else "",
            "regiao": row["regiao"],
            "area": row["area"],
            "quarteirao": row["quarteirao"],
            "logradouro": row["logradouro"],
            "imovel": row["imovel"],
            "data_chegada": row["data_chegada"],
            "data_saida": row["data_saida"],
            "motivo_duplicacao": row["motivo_duplicacao"],
            "_fill": "FFD9B3",  # laranja claro para destacar
        })
    return registros


def _preparar_suspeitas(df_total):
    """
    Monta a lista de visitas marcadas como 'sequência suspeita' (chegadas no
    mesmo minuto que a visita anterior do mesmo agente), com endereço completo
    e o número da cadeia a que pertencem, para facilitar a checagem manual.
    """
    if df_total.empty:
        return []

    susp = df_total[df_total["sequencia_suspeita"]].copy()
    if susp.empty:
        return []

    # Ordena por agente e depois pela cadeia/horário, para que cada grupo
    # suspeito apareça junto, na ordem em que ocorreu.
    susp = susp.sort_values(["nome_agente", "dt_chegada"])

    registros = []
    for _, row in susp.iterrows():
        # Só exibe a diferença quando ELA é a responsável pelo alerta desta
        # linha (par com a visita ANTERIOR). Se o alerta veio do par com a
        # visita seguinte, o diff em relação à anterior é irrelevante/enganoso.
        diff_seg = row["diff_chegada_anterior_seg"] if row["par_suspeito_anterior"] else None
        registros.append({
            "nome_agente": row["nome_agente"],
            "equipe": row["equipe"],
            "grupo_sequencia_suspeita": row["grupo_sequencia_suspeita"],
            "id_visita": row["id_visita"] if pd.notna(row["id_visita"]) else "",
            "regiao": row["regiao"],
            "area": row["area"],
            "quarteirao": row["quarteirao"],
            "logradouro": row["logradouro"],
            "imovel": row["imovel"],
            "data_chegada": row["data_chegada"],
            "data_saida": row["data_saida"],
            "diff_chegada_anterior_seg": "" if diff_seg is None or pd.isna(diff_seg) else int(diff_seg),
            "_fill": "FFE0B3",  # amarelo/laranja claro para destacar
        })
    return registros


def _preparar_negativas(df_total):
    """
    Monta a lista de visitas com duração NEGATIVA (data de saída registrada
    antes da chegada) — provável bug de gravação no sistema. Traz endereço
    completo para facilitar a investigação junto ao suporte do e-Visita.
    """
    if df_total.empty:
        return []

    neg = df_total[df_total["alerta_neg"]].copy()
    if neg.empty:
        return []

    neg = neg.sort_values(["nome_agente", "dt_chegada"])

    registros = []
    for _, row in neg.iterrows():
        registros.append({
            "nome_agente": row["nome_agente"],
            "equipe": row["equipe"],
            "id_visita": row["id_visita"] if pd.notna(row["id_visita"]) else "",
            "regiao": row["regiao"],
            "area": row["area"],
            "quarteirao": row["quarteirao"],
            "logradouro": row["logradouro"],
            "imovel": row["imovel"],
            "data_chegada": row["data_chegada"],
            "data_saida": row["data_saida"],
            "duracao_min": row["duracao_min"],
            "observacao": "Data/hora de saída anterior à chegada — provável erro de gravação no sistema",
            "_fill": "D9D9D9",  # cinza para destacar como problema técnico, não de conduta
        })
    return registros


def _preparar_acima_teto_media(df_total):
    """
    Monta a lista de visitas com duração acima de TETO_DURACAO_MEDIA_MIN
    (15min por padrão) — pra auditoria de possível manipulação de horário
    de saída (lançamento pelo computador). Essas visitas entram no cálculo
    de média/mediana com duração limitada ao teto, mas aqui aparecem com a
    duração REAL registrada, pra facilitar a checagem.
    """
    if df_total.empty:
        return []

    acima = df_total[df_total["acima_teto_media"]].copy()
    if acima.empty:
        return []

    acima = acima.sort_values("duracao_min", ascending=False)

    registros = []
    for _, row in acima.iterrows():
        registros.append({
            "nome_agente": row["nome_agente"],
            "equipe": row["equipe"],
            "id_visita": row["id_visita"] if pd.notna(row["id_visita"]) else "",
            "regiao": row["regiao"],
            "area": row["area"],
            "quarteirao": row["quarteirao"],
            "logradouro": row["logradouro"],
            "imovel": row["imovel"],
            "data_chegada": row["data_chegada"],
            "data_saida": row["data_saida"],
            "duracao_min": row["duracao_min"],
        })
    return registros


def _motivo_fora_expediente(row):
    if row["fim_de_semana"]:
        return f"Fim de semana ({NOMES_DIA_SEMANA.get(row['dia_semana_num'], '')})"
    minutos = row["minutos_chegada"]
    hora = row["hora_chegada"]
    if HORA_ALMOCO_INICIO_MIN <= minutos < HORA_ALMOCO_FIM_MIN:
        return f"Horário de almoço ({_fmt_hora_min(HORA_ALMOCO_INICIO_MIN)}-{_fmt_hora_min(HORA_ALMOCO_FIM_MIN)})"
    if hora < HORA_EXPEDIENTE_MANHA_INICIO:
        return f"Antes do expediente (antes das {HORA_EXPEDIENTE_MANHA_INICIO}h)"
    if hora >= HORA_EXPEDIENTE_TARDE_FIM:
        return f"Depois do expediente (após as {HORA_EXPEDIENTE_TARDE_FIM}h)"
    return "Fora do expediente (próximo ao horário de almoço, dentro da tolerância de 15min)"


def _preparar_fora_expediente(df_total):
    """
    Monta a lista de visitas registradas fora do horário oficial de serviço
    (7h-11h e 13h-17h, dias úteis) — inclui fins de semana, antes/depois do
    expediente e horário de almoço (com motivo especificado por linha).
    """
    if df_total.empty:
        return []

    fora = df_total[df_total["fora_expediente"]].copy()
    if fora.empty:
        return []

    fora["motivo"] = fora.apply(_motivo_fora_expediente, axis=1)
    fora["dia_semana"] = fora["dia_semana_num"].map(NOMES_DIA_SEMANA)
    fora = fora.sort_values(["nome_agente", "dt_chegada"])

    registros = []
    for _, row in fora.iterrows():
        registros.append({
            "nome_agente": row["nome_agente"],
            "equipe": row["equipe"],
            "id_visita": row["id_visita"] if pd.notna(row["id_visita"]) else "",
            "dia_semana": row["dia_semana"],
            "regiao": row["regiao"],
            "area": row["area"],
            "quarteirao": row["quarteirao"],
            "logradouro": row["logradouro"],
            "imovel": row["imovel"],
            "data_chegada": row["data_chegada"],
            "data_saida": row["data_saida"],
            "motivo": row["motivo"],
            "_fill": "FFCCCC" if row["fim_de_semana"] else "FFE8CC",
        })
    return registros


def _preparar_almoco(df_total):
    """Lista apenas as visitas registradas no horário de almoço (11h15-12h45, com tolerância de 15min)."""
    if df_total.empty:
        return []

    alm = df_total[df_total["visita_horario_almoco"]].copy()
    if alm.empty:
        return []

    alm = alm.sort_values(["nome_agente", "dt_chegada"])

    registros = []
    for _, row in alm.iterrows():
        registros.append({
            "nome_agente": row["nome_agente"],
            "equipe": row["equipe"],
            "id_visita": row["id_visita"] if pd.notna(row["id_visita"]) else "",
            "regiao": row["regiao"],
            "area": row["area"],
            "quarteirao": row["quarteirao"],
            "logradouro": row["logradouro"],
            "imovel": row["imovel"],
            "data_chegada": row["data_chegada"],
            "data_saida": row["data_saida"],
            "_fill": "FFF2CC",
        })
    return registros


def _preparar_ausencias(df_total, resultados):
    """
    Para cada agente, dia útil e turno (Manhã 7h-11h / Tarde 13h-17h) dentro
    do período coberto pelos dados coletados, verifica se houve pelo menos
    uma visita registrada. Se não houve, lista como possível ausência
    (folga, atestado, falta — precisa confirmar com a escala da equipe).

    O período considerado é do primeiro ao último dia com QUALQUER visita
    coletada (de qualquer agente), já que não temos o calendário oficial de
    escala/férias de cada um.
    """
    if df_total.empty:
        return []

    dia_min = df_total["dia"].min()
    dia_max = df_total["dia"].max()
    dias_uteis = pd.bdate_range(start=dia_min, end=dia_max)  # só dias úteis (seg-sex)

    turnos_oficiais = [
        ("Manhã", HORA_EXPEDIENTE_MANHA_INICIO, HORA_EXPEDIENTE_MANHA_FIM),
        ("Tarde", HORA_EXPEDIENTE_TARDE_INICIO, HORA_EXPEDIENTE_TARDE_FIM),
    ]

    registros = []
    for res in resultados:
        df_ag = df_total[df_total["id_agente"] == res["id_agente"]]
        dias_com_visita = set(df_ag["dia"])

        for dia_ts in dias_uteis:
            dia = dia_ts.date()
            if dia not in dias_com_visita:
                # Nenhuma visita nesse dia inteiro: ausência nos dois turnos
                turnos_ausentes = turnos_oficiais
            else:
                sub_dia = df_ag[df_ag["dia"] == dia]
                turnos_ausentes = [
                    (nome, ini, fim) for nome, ini, fim in turnos_oficiais
                    if not ((sub_dia["hora_chegada"] >= ini) & (sub_dia["hora_chegada"] < fim)).any()
                ]

            for turno_nome, _, _ in turnos_ausentes:
                motivo = _obter_motivo_cronograma(res["id_agente"], res["equipe"], dia)
                if motivo:
                    continue  # dia/turno já explicado (cronograma_ausencias.json) — não é ausência a investigar
                registros.append({
                    "nome_agente": res["nome_agente"],
                    "equipe": res["equipe"],
                    "dia_fmt": dia.strftime("%d/%m/%Y"),
                    "dia_semana": NOMES_DIA_SEMANA.get(dia_ts.weekday(), ""),
                    "turno": turno_nome,
                    "observacao": "Sem visita registrada nesse turno — possível folga, atestado ou falta (confirmar com a escala)",
                    "_fill": "E8E8E8",
                })

    # Detecta ausência COLETIVA: se TODOS os agentes de uma equipe (entre os
    # que aparecem nos dados) faltaram no mesmo dia/turno, é mais provável
    # que seja um motivo coletivo (reunião de equipe, chuva/intempérie,
    # feriado local, evento) do que folga/atestado individual coincidente.
    total_agentes_por_equipe = {}
    for res in resultados:
        total_agentes_por_equipe.setdefault(res["equipe"], set()).add(res["id_agente"])
    total_agentes_por_equipe = {eq: len(ids) for eq, ids in total_agentes_por_equipe.items()}

    contagem_ausentes = {}
    for r in registros:
        chave = (r["equipe"], r["dia_fmt"], r["turno"])
        contagem_ausentes[chave] = contagem_ausentes.get(chave, 0) + 1

    for r in registros:
        chave = (r["equipe"], r["dia_fmt"], r["turno"])
        total_equipe = total_agentes_por_equipe.get(r["equipe"], 0)
        if total_equipe >= 2 and contagem_ausentes[chave] == total_equipe:
            r["observacao"] = (f"TODA A EQUIPE ({total_equipe} agentes) ausente nesse turno — "
                                 f"possível reunião de equipe, chuva/intempérie ou evento — verificar")
            r["_fill"] = "FFD9A0"

    registros.sort(key=lambda r: (r["equipe"], r["dia_fmt"], r["turno"], r["nome_agente"]))
    return registros


def _preparar_cronograma(resultados, data_min=None, data_max=None):
    """
    Lista, de forma unificada e ordenada por data, todo o cronograma de
    ausências programadas (CRONOGRAMA_AUSENCIAS — chuva, folga, atestado,
    reunião/treinamento, férias, levantamento de índice, ovitrampa,
    recuperação de casas — alvo agente/equipe/todos).

    Se data_min/data_max forem passados (o período coberto pela planilha
    sendo gerada, tirado de df_total), só entram registros cujo período se
    sobrepõe a esse intervalo — é isso que faz o Cronograma ficar
    INDEPENDENTE por mês: gerando Agosto, só aparecem os registros de
    Agosto; gerando Julho, só os de Julho, mesmo que
    data/cronograma_ausencias.json tenha entradas de vários meses juntas.
    """
    nomes_por_id = {res["id_agente"]: res["nome_agente"] for res in resultados}

    registros = []
    for entrada in CRONOGRAMA_AUSENCIAS:
        try:
            entrada_inicio = pd.to_datetime(entrada.get("inicio", ""), format="%d/%m/%Y", errors="coerce")
            entrada_fim = pd.to_datetime(entrada.get("fim", ""), format="%d/%m/%Y", errors="coerce")
        except (ValueError, TypeError):
            entrada_inicio = entrada_fim = pd.NaT
        if data_min is not None and data_max is not None and pd.notna(entrada_inicio) and pd.notna(entrada_fim):
            # Sobreposição de intervalos: sobra fora se o registro termina
            # antes do período começar OU começa depois do período terminar.
            if entrada_fim.date() < data_min or entrada_inicio.date() > data_max:
                continue

        alvo = entrada.get("alvo", "agente")
        if alvo == "todos":
            alvo_fmt = "Todos os agentes"
        elif alvo == "equipe":
            alvo_fmt = entrada.get("equipe", "—")
        else:
            id_ag = entrada.get("id_agente")
            alvo_fmt = nomes_por_id.get(id_ag, f"Agente_{id_ag}")
        registros.append({
            "dia_inicio": entrada.get("inicio", ""),
            "dia_fim": entrada.get("fim", ""),
            "motivo": entrada.get("motivo", ""),
            "alvo_fmt": alvo_fmt,
            "_fill": {"Chuva": "D6E4F0", "Reunião/Treinamento": "D6E4F0",
                      "Férias": "E0D6F0", "Folga": "E0D6F0", "Atestado": "F0DCD6",
                      "Levantamento de Índice": "D6F0DC", "Instalação de Ovitrampa": "D6F0DC",
                      "Recolha de Ovitrampa": "D6F0DC", "Recuperação de Casas": "D6F0DC"}.get(entrada.get("motivo"), "E8E8E8"),
            "_ordenacao": entrada_inicio,
        })

    registros.sort(key=lambda r: r["_ordenacao"] or pd.Timestamp.min)
    for r in registros:
        del r["_ordenacao"]
    return registros


def _preparar_pendencias_resumo(resultados):
    """
    Resumo de pendência por agente (abertos/fechados/recusados/total/% pendência),
    com uma linha de TOTAL GERAL consolidando todos os agentes ao final.
    """
    registros = []
    soma_abertos = soma_fechados = soma_recusados = 0
    for res in resultados:
        r = res["resumo"]
        abertos = r["total_visitas"]
        fechados = r.get("imoveis_fechados", 0)
        recusados = r.get("imoveis_recusados", 0)
        total_geral = r.get("total_geral_pendencia", abertos + fechados + recusados)
        pct = r.get("pct_pendencia", 0)

        soma_abertos += abertos
        soma_fechados += fechados
        soma_recusados += recusados

        registros.append({
            "nome_agente": res["nome_agente"],
            "equipe": res["equipe"],
            "abertos": abertos,
            "fechados": fechados,
            "recusados": recusados,
            "total_geral": total_geral,
            "pct_pendencia": pct,
        })

    soma_geral = soma_abertos + soma_fechados + soma_recusados
    pct_geral = round(((soma_fechados + soma_recusados) / soma_geral * 100), 2) if soma_geral > 0 else 0
    registros.append({
        "nome_agente": "TOTAL GERAL",
        "equipe": "—",
        "abertos": soma_abertos,
        "fechados": soma_fechados,
        "recusados": soma_recusados,
        "total_geral": soma_geral,
        "pct_pendencia": pct_geral,
        "_fill": "D9D9D9",
    })
    return registros


def _preparar_ranking(resultados, regs_ausencias):
    """
    Calcula a pontuação de cada agente (calcular_pontuacao) e monta:
      - lista principal ordenada por pontuação (posição 1 = maior pontuação)
      - lista de detalhamento (todos os critérios aplicados, por agente)
    """
    ausencias_por_agente = {}
    for r in regs_ausencias:
        ausencias_por_agente[r["nome_agente"]] = ausencias_por_agente.get(r["nome_agente"], 0) + 1

    principal, todos_detalhes = [], []
    for res in resultados:
        qtd_ausencias = ausencias_por_agente.get(res["nome_agente"], 0)
        pontos, classif, fill, detalhes = calcular_pontuacao(res, qtd_ausencias)
        principal.append({
            "nome_agente": res["nome_agente"],
            "equipe": res["equipe"],
            "pontos": pontos,
            "classificacao": classif,
            "_fill": fill,
        })
        for d in detalhes:
            d["_fill"] = "FFE0E0" if d["pontos"] < 0 else "E0FFE0"
        todos_detalhes.extend(detalhes)

    principal.sort(key=lambda r: r["pontos"], reverse=True)
    for i, r in enumerate(principal, 1):
        r["posicao"] = i

    todos_detalhes.sort(key=lambda d: (d["nome_agente"], d["motivo"]))
    return principal, todos_detalhes


def _preparar_recusados_detalhe(resultados):
    """Lista individual de todos os imóveis com visita recusada pelo morador."""
    registros = []
    for res in resultados:
        for v in res.get("visitas_recusados_raw", []):
            registros.append({
                "nome_agente": v["nome_agente"],
                "equipe": v["equipe"],
                "id_visita": v["id_visita"] if v.get("id_visita") else "",
                "regiao": v["regiao"],
                "area": v["area"],
                "quarteirao": v["quarteirao"],
                "logradouro": v["logradouro"],
                "imovel": v["imovel"],
                "data_chegada": v["data_chegada"],
                "data_saida": v["data_saida"],
                "_fill": "F4CCCC",
            })
    registros.sort(key=lambda r: (r["nome_agente"], r["data_chegada"]))
    return registros


def _slugify_nome_arquivo(nome):
    """Converte o título de uma aba (com espaços/acentos) num nome de arquivo seguro."""
    sem_acento = unicodedata.normalize("NFKD", nome).encode("ascii", "ignore").decode("ascii")
    limpo = re.sub(r"[^\w\s-]", "", sem_acento).strip().replace(" ", "_")
    return limpo or "aba"


def exportar_todas_abas_csv(wb, pasta_saida):
    """
    Exporta CADA aba do workbook como um arquivo .csv separado (com os
    mesmos valores exibidos no Excel), já que um único CSV não suporta
    múltiplas abas. Os arquivos ficam em <pasta_saida>/csv/.
    """
    pasta_csv = os.path.join(pasta_saida, "csv")
    os.makedirs(pasta_csv, exist_ok=True)
    arquivos_gerados = []

    for ws in wb.worksheets:
        nome_arquivo = _slugify_nome_arquivo(ws.title) + ".csv"
        caminho_csv = os.path.join(pasta_csv, nome_arquivo)
        try:
            with open(caminho_csv, "w", newline="", encoding="utf-8-sig") as f:
                escritor = csv.writer(f, delimiter=";")
                for row in ws.iter_rows(values_only=True):
                    escritor.writerow(["" if v is None else v for v in row])
            arquivos_gerados.append(caminho_csv)
        except OSError as e:
            log.warning("Não foi possível exportar a aba '%s' para CSV: %s", ws.title, e)

    log.info("✅ %s abas exportadas em CSV, em: %s", len(arquivos_gerados), pasta_csv)
    return pasta_csv, arquivos_gerados


def salvar_excel_consolidado(resultados, df_total, pasta_saida, semanas_ponto_estrategico=None,
                              semanas_tratamento=None):
    os.makedirs(pasta_saida, exist_ok=True)
    caminho = os.path.join(pasta_saida, ARQUIVO_CONSOLIDADO)
    log.info("💾 Salvando %s", caminho)

    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Resumo Geral")
    for col, (titulo, _) in enumerate(COLS_RESUMO, 1):
        ws.cell(row=2, column=col, value=titulo)
    registros_resumo = [_preparar_resumo(r) for r in resultados]
    linha_final = _escrever(ws, 3, registros_resumo, COLS_RESUMO)
    ultima_linha_dados = linha_final - 1
    _estilizar(ws, 2, titulo="RESUMO CONSOLIDADO — TODOS OS AGENTES", ultima_linha_dados=ultima_linha_dados)
    _adicionar_legenda(ws, linha_final + 2)

    ws_a = wb.create_sheet("Por Área")
    for col, (titulo, _) in enumerate(COLS_AREA, 1):
        ws_a.cell(row=2, column=col, value=titulo)
    regs = []
    for res in resultados:
        regs.extend(_preparar_area(df_total[df_total["id_agente"] == res["id_agente"]], res))
    linha_final_a = _escrever(ws_a, 3, regs, COLS_AREA)
    ultima_linha_dados_a = linha_final_a - 1
    _estilizar(ws_a, 2, cor_cab="1F3864", titulo="ANÁLISE POR ÁREA", ultima_linha_dados=ultima_linha_dados_a)
    _adicionar_legenda(ws_a, linha_final_a + 2)

    ws_d = wb.create_sheet("Visitas Duplicadas")
    for col, (titulo, _) in enumerate(COLS_DUPLICADAS, 1):
        ws_d.cell(row=2, column=col, value=titulo)
    regs_dup = _preparar_duplicadas(df_total)
    linha_final_d = _escrever(ws_d, 3, regs_dup, COLS_DUPLICADAS)
    ultima_linha_dados_d = linha_final_d - 1
    _estilizar(ws_d, 2, cor_cab="8C3B00",
               titulo=f"VISITAS DUPLICADAS ({len(regs_dup)} registros) — CONFIRME E EXCLUA NO SISTEMA",
               ultima_linha_dados=ultima_linha_dados_d)

    ws_s = wb.create_sheet("Visitas Suspeitas")
    for col, (titulo, _) in enumerate(COLS_SUSPEITAS, 1):
        ws_s.cell(row=2, column=col, value=titulo)
    regs_susp = _preparar_suspeitas(df_total)
    linha_final_s = _escrever(ws_s, 3, regs_susp, COLS_SUSPEITAS)
    ultima_linha_dados_s = linha_final_s - 1
    _estilizar(ws_s, 2, cor_cab="7A4B00",
               titulo=f"VISITAS EM SEQUÊNCIA SUSPEITA ({len(regs_susp)} registros) — "
                      f"CHEGADA NO MESMO MINUTO DA VISITA ANTERIOR",
               ultima_linha_dados=ultima_linha_dados_s)
    _adicionar_legenda(ws_s, linha_final_s + 2)

    ws_n = wb.create_sheet("Visitas Negativas")
    for col, (titulo, _) in enumerate(COLS_NEGATIVAS, 1):
        ws_n.cell(row=2, column=col, value=titulo)
    regs_neg = _preparar_negativas(df_total)
    linha_final_n = _escrever(ws_n, 3, regs_neg, COLS_NEGATIVAS)
    ultima_linha_dados_n = linha_final_n - 1
    _estilizar(ws_n, 2, cor_cab="595959",
               titulo=f"VISITAS COM DURAÇÃO NEGATIVA ({len(regs_neg)} registros) — "
                      f"PROVÁVEL BUG DO SISTEMA, EXCLUÍDAS DA MÉDIA/MEDIANA",
               ultima_linha_dados=ultima_linha_dados_n)
    _adicionar_legenda(ws_n, linha_final_n + 2)

    ws_teto = wb.create_sheet("Visitas Acima de 15min")
    for col, (titulo, _) in enumerate(COLS_ACIMA_TETO_MEDIA, 1):
        ws_teto.cell(row=2, column=col, value=titulo)
    regs_teto = _preparar_acima_teto_media(df_total)
    linha_final_teto = _escrever(ws_teto, 3, regs_teto, COLS_ACIMA_TETO_MEDIA)
    ultima_linha_dados_teto = linha_final_teto - 1
    _estilizar(ws_teto, 2, cor_cab="4A5568",
               titulo=f"VISITAS COM DURAÇÃO ACIMA DE {TETO_DURACAO_MEDIA_MIN}MIN ({len(regs_teto)} registros) — "
                      f"AUDITORIA DE POSSÍVEL MANIPULAÇÃO DE HORÁRIO (LANÇAMENTO PELO COMPUTADOR); "
                      f"NO CÁLCULO DE MÉDIA/MEDIANA ENTRAM COMO {TETO_DURACAO_MEDIA_MIN}MIN",
               ultima_linha_dados=ultima_linha_dados_teto)

    ws_fe = wb.create_sheet("Fora do Expediente")
    for col, (titulo, _) in enumerate(COLS_FORA_EXPEDIENTE, 1):
        ws_fe.cell(row=2, column=col, value=titulo)
    regs_fe = _preparar_fora_expediente(df_total)
    linha_final_fe = _escrever(ws_fe, 3, regs_fe, COLS_FORA_EXPEDIENTE)
    ultima_linha_dados_fe = linha_final_fe - 1
    _estilizar(ws_fe, 2, cor_cab="8C1F1F",
               titulo=f"VISITAS FORA DO EXPEDIENTE / FIM DE SEMANA ({len(regs_fe)} registros) — "
                      f"HORÁRIO OFICIAL: {HORA_EXPEDIENTE_MANHA_INICIO}h-{HORA_EXPEDIENTE_MANHA_FIM}h "
                      f"E {HORA_EXPEDIENTE_TARDE_INICIO}h-{HORA_EXPEDIENTE_TARDE_FIM}h, DIAS ÚTEIS",
               ultima_linha_dados=ultima_linha_dados_fe)

    ws_al = wb.create_sheet("Visitas no Almoço")
    for col, (titulo, _) in enumerate(COLS_ALMOCO, 1):
        ws_al.cell(row=2, column=col, value=titulo)
    regs_al = _preparar_almoco(df_total)
    linha_final_al = _escrever(ws_al, 3, regs_al, COLS_ALMOCO)
    ultima_linha_dados_al = linha_final_al - 1
    _estilizar(ws_al, 2, cor_cab="1F6F5C",
               titulo=f"VISITAS NO HORÁRIO DE ALMOÇO ({len(regs_al)} registros) — "
                      f"{_fmt_hora_min(HORA_ALMOCO_INICIO_MIN)} ÀS {_fmt_hora_min(HORA_ALMOCO_FIM_MIN)} "
                      f"(TOLERÂNCIA DE 15MIN JÁ APLICADA)",
               ultima_linha_dados=ultima_linha_dados_al)

    ws_aus = wb.create_sheet("Ausências")
    for col, (titulo, _) in enumerate(COLS_AUSENCIAS, 1):
        ws_aus.cell(row=2, column=col, value=titulo)
    regs_aus = _preparar_ausencias(df_total, resultados)
    linha_final_aus = _escrever(ws_aus, 3, regs_aus, COLS_AUSENCIAS)
    ultima_linha_dados_aus = linha_final_aus - 1
    _estilizar(ws_aus, 2, cor_cab="4A4A4A",
               titulo=f"POSSÍVEIS AUSÊNCIAS POR DIA/TURNO ({len(regs_aus)} registros) — "
                      f"CONFIRMAR COM A ESCALA (FOLGA/ATESTADO/FALTA)",
               ultima_linha_dados=ultima_linha_dados_aus)

    ws_cron = wb.create_sheet("Cronograma")
    for col, (titulo, _) in enumerate(COLS_CRONOGRAMA, 1):
        ws_cron.cell(row=2, column=col, value=titulo)
    periodo_dias = df_total["dia"] if "dia" in df_total.columns and not df_total.empty else None
    data_min = periodo_dias.min() if periodo_dias is not None and not periodo_dias.empty else None
    data_max = periodo_dias.max() if periodo_dias is not None and not periodo_dias.empty else None
    regs_cron = _preparar_cronograma(resultados, data_min, data_max)
    linha_final_cron = _escrever(ws_cron, 3, regs_cron, COLS_CRONOGRAMA)
    ultima_linha_dados_cron = linha_final_cron - 1
    _estilizar(ws_cron, 2, cor_cab="2E5C8A",
               titulo=f"CRONOGRAMA — CHUVA/FOLGA/ATESTADO/REUNIÃO/FÉRIAS ({len(regs_cron)} registros) — "
                      f"EXCLUÍDOS DA ABA 'AUSÊNCIAS'. Edite em data/cronograma_ausencias.json "
                      f"(use cronograma_editor.html).",
               ultima_linha_dados=ultima_linha_dados_cron)

    ws_pend = wb.create_sheet("Recusados e Pendências")
    for col, (titulo, _) in enumerate(COLS_PENDENCIAS_RESUMO, 1):
        ws_pend.cell(row=2, column=col, value=titulo)
    regs_pend = _preparar_pendencias_resumo(resultados)
    linha_final_pend = _escrever(ws_pend, 3, regs_pend, COLS_PENDENCIAS_RESUMO)
    _estilizar(ws_pend, 2, cor_cab="7A1F1F",
               titulo="RESUMO DE PENDÊNCIA POR AGENTE (com Total Geral) — "
                      "% Pendência = (Fechados + Recusados) / (Abertos+Fechados+Recusados)")

    linha_detalhe_inicio = linha_final_pend + 2
    ws_pend.cell(row=linha_detalhe_inicio, column=1,
                  value=f"DETALHE — IMÓVEIS COM VISITA RECUSADA PELO MORADOR")
    ws_pend.cell(row=linha_detalhe_inicio, column=1).font = Font(bold=True, size=11, color="FFFFFF")
    ws_pend.cell(row=linha_detalhe_inicio, column=1).fill = PatternFill("solid", fgColor="7A1F1F")
    ws_pend.merge_cells(start_row=linha_detalhe_inicio, start_column=1,
                          end_row=linha_detalhe_inicio, end_column=len(COLS_RECUSADOS_DETALHE))
    for col, (titulo, _) in enumerate(COLS_RECUSADOS_DETALHE, 1):
        cell = ws_pend.cell(row=linha_detalhe_inicio + 1, column=col, value=titulo)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor="A34A4A")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    regs_recusados = _preparar_recusados_detalhe(resultados)
    linha_final_recusados = _escrever(ws_pend, linha_detalhe_inicio + 2, regs_recusados, COLS_RECUSADOS_DETALHE)

    # Sem autofiltro aqui: são duas tabelas com colunas diferentes empilhadas
    # na mesma aba, então um filtro único não faria sentido. Só recalcula a
    # largura das colunas considerando todo o conteúdo já escrito.
    _estilizar(ws_pend, 2)

    if semanas_ponto_estrategico:
        ws_pe = wb.create_sheet("Ponto Estratégico")
        for col, (titulo, _) in enumerate(COLS_PONTO_ESTRATEGICO, 1):
            ws_pe.cell(row=2, column=col, value=titulo)
        regs_pe = _preparar_ponto_estrategico(semanas_ponto_estrategico)
        linha_final_pe = _escrever(ws_pe, 3, regs_pe, COLS_PONTO_ESTRATEGICO)
        ultima_linha_dados_pe = linha_final_pe - 1
        _estilizar(ws_pe, 2, cor_cab="4A2E6B",
                   titulo=f"PONTO ESTRATÉGICO — RESULTADOS POR SEMANA ({len(regs_pe)} linhas)",
                   ultima_linha_dados=ultima_linha_dados_pe)
        _adicionar_legenda(ws_pe, linha_final_pe + 2)

    # Resumo Semanal (tratamento) — mesma métrica do Resumo Geral, só que
    # quebrada por semana em vez de somada no período inteiro. Alimenta o
    # filtro "por semana" do dashboard (aba Resumo/Agentes). Só existe
    # quando o período consolidado cobre mais de uma semana — com uma
    # semana só ela seria idêntica ao Resumo Geral.
    if semanas_tratamento:
        ws_rs = wb.create_sheet("Resumo Semanal")
        for col, (titulo, _) in enumerate(COLS_PONTO_ESTRATEGICO, 1):
            ws_rs.cell(row=2, column=col, value=titulo)
        regs_rs = _preparar_ponto_estrategico(semanas_tratamento)
        linha_final_rs = _escrever(ws_rs, 3, regs_rs, COLS_PONTO_ESTRATEGICO)
        ultima_linha_dados_rs = linha_final_rs - 1
        _estilizar(ws_rs, 2, cor_cab="0F5132",
                   titulo=f"RESUMO GERAL — QUEBRADO POR SEMANA ({len(regs_rs)} linhas)",
                   ultima_linha_dados=ultima_linha_dados_rs)
        _adicionar_legenda(ws_rs, linha_final_rs + 2)

    # --- Ranking / Pontuação ---
    ranking_principal, ranking_detalhe = _preparar_ranking(resultados, regs_aus)

    ws_rk = wb.create_sheet("Ranking")
    for col, (titulo, _) in enumerate(COLS_RANKING, 1):
        ws_rk.cell(row=2, column=col, value=titulo)
    linha_final_rk = _escrever(ws_rk, 3, ranking_principal, COLS_RANKING)
    ultima_linha_dados_rk = linha_final_rk - 1
    _estilizar(ws_rk, 2, cor_cab="1A1A2E",
               titulo=f"RANKING DOS AGENTES — TODOS COMEÇAM COM {PONTOS_INICIAIS} PONTOS",
               ultima_linha_dados=ultima_linha_dados_rk)

    linha_detalhe_rk_inicio = linha_final_rk + 2
    ws_rk.cell(row=linha_detalhe_rk_inicio, column=1,
                value="DETALHAMENTO — TODOS OS CRITÉRIOS APLICADOS POR AGENTE")
    ws_rk.cell(row=linha_detalhe_rk_inicio, column=1).font = Font(bold=True, size=11, color="FFFFFF")
    ws_rk.cell(row=linha_detalhe_rk_inicio, column=1).fill = PatternFill("solid", fgColor="1A1A2E")
    ws_rk.merge_cells(start_row=linha_detalhe_rk_inicio, start_column=1,
                        end_row=linha_detalhe_rk_inicio, end_column=len(COLS_RANKING_DETALHE))
    for col, (titulo, _) in enumerate(COLS_RANKING_DETALHE, 1):
        cell = ws_rk.cell(row=linha_detalhe_rk_inicio + 1, column=col, value=titulo)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor="2E4057")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    _escrever(ws_rk, linha_detalhe_rk_inicio + 2, ranking_detalhe, COLS_RANKING_DETALHE)

    # Sem autofiltro: duas tabelas com colunas diferentes empilhadas na mesma aba.
    _estilizar(ws_rk, 2)

    wb.save(caminho)
    log.info("✅ Consolidado salvo!")

    # Exporta TODAS as abas também em CSV (arquivo idêntico ao Excel, só que
    # em CSV — como CSV não suporta múltiplas abas, gera um .csv por aba).
    exportar_todas_abas_csv(wb, pasta_saida)

    return caminho


# ============================================================
# ORQUESTRAÇÃO
# ============================================================

def processar(equipe, id_ag, session_base, cfg):
    session = clonar_session(session_base)
    try:
        nome, visitas, paginas_falhas = coletar_visitas_agente(session, id_ag, cfg,
                                                                  id_pendencia=ID_PENDENCIA_ABERTOS)
    except SessaoExpiradaError as e:
        log.error("Sessão expirou processando agente %s: %s", id_ag, e)
        return {"ok": False, "id_agente": id_ag, "nome_agente": f"Agente_{id_ag}",
                "equipe": equipe, "erro": "sessao_expirada"}
    except Exception as e:
        log.error("❌ Falha ao coletar página 1 do agente %s mesmo após %s tentativas: %s — "
                    "AGENTE FICARÁ AUSENTE DO RELATÓRIO", id_ag, TENTATIVAS_POR_PAGINA, e)
        return {"ok": False, "id_agente": id_ag, "nome_agente": f"Agente_{id_ag}",
                "equipe": equipe, "erro": "falha_pagina1"}

    if not visitas:
        return {"ok": False, "id_agente": id_ag, "nome_agente": nome, "equipe": equipe}

    # Coleta adicional de fechados/recusados, só para contagem de pendência
    # (não entra na análise de tempo/qualidade, que é sempre sobre os abertos).
    visitas_fechados, visitas_recusados = [], []
    paginas_falhas_pendencia = []
    if COLETAR_FECHADOS_RECUSADOS:
        try:
            _, visitas_fechados, pf_fechados = coletar_visitas_agente(
                session, id_ag, cfg, id_pendencia=ID_PENDENCIA_FECHADOS, obter_nome=False)
            paginas_falhas_pendencia.extend(pf_fechados)
        except SessaoExpiradaError:
            raise
        except Exception as e:
            log.warning("Não foi possível coletar FECHADOS do agente %s: %s", id_ag, e)

        try:
            _, visitas_recusados, pf_recusados = coletar_visitas_agente(
                session, id_ag, cfg, id_pendencia=ID_PENDENCIA_RECUSADOS, obter_nome=False)
            paginas_falhas_pendencia.extend(pf_recusados)
        except SessaoExpiradaError:
            raise
        except Exception as e:
            log.warning("Não foi possível coletar RECUSADOS do agente %s: %s", id_ag, e)

    df = calcular_tempos(visitas, nome, id_ag, equipe, semana_html_id=cfg["semana_inicio"])
    resumo = resumo_agente(df)

    total_abertos = resumo["total_visitas"]
    total_fechados = len(visitas_fechados)
    total_recusados = len(visitas_recusados)
    total_geral_pendencia = total_abertos + total_fechados + total_recusados
    pct_pendencia = round(((total_fechados + total_recusados) / total_geral_pendencia * 100), 2) if total_geral_pendencia > 0 else 0

    resumo["imoveis_fechados"] = total_fechados
    resumo["imoveis_recusados"] = total_recusados
    resumo["total_geral_pendencia"] = total_geral_pendencia
    resumo["pct_pendencia"] = pct_pendencia

    # Contagem por área (só contagem — não precisa de análise de tempo/qualidade)
    fechados_por_area = {}
    for v in visitas_fechados:
        fechados_por_area[v["area"]] = fechados_por_area.get(v["area"], 0) + 1
    recusados_por_area = {}
    for v in visitas_recusados:
        recusados_por_area[v["area"]] = recusados_por_area.get(v["area"], 0) + 1

    visitas_recusados_raw = [{**v, "nome_agente": nome, "equipe": equipe} for v in visitas_recusados]

    return {
        "ok": True,
        "id_agente": id_ag,
        "nome_agente": nome,
        "equipe": equipe,
        "resumo": resumo,
        "df": df,
        "paginas_falhas": paginas_falhas + paginas_falhas_pendencia,
        "fechados_por_area": fechados_por_area,
        "recusados_por_area": recusados_por_area,
        "visitas_recusados_raw": visitas_recusados_raw,
    }


def processar_ponto_estrategico(id_ag, session_base, cfg_pe):
    """
    Coleta as visitas de Ponto Estratégico de um agente (URL/id_atividade
    próprios) e quebra o resultado POR SEMANA, calculando o mesmo conjunto
    completo de métricas do Resumo Geral (reaproveita resumo_agente) para
    cada semana separadamente.
    """
    session = clonar_session(session_base)
    try:
        nome, visitas, paginas_falhas = coletar_visitas_agente(
            session, id_ag, cfg_pe, id_pendencia=ID_PENDENCIA_ABERTOS,
            id_atividade=ID_ATIVIDADE_PONTO_ESTRATEGICO, tem_foto="")
    except SessaoExpiradaError as e:
        log.error("Sessão expirou processando Ponto Estratégico do agente %s: %s", id_ag, e)
        return {"ok": False, "id_agente": id_ag, "erro": "sessao_expirada"}
    except Exception as e:
        log.error("❌ Falha ao coletar Ponto Estratégico do agente %s: %s — "
                    "AGENTE FICARÁ AUSENTE DESSA ABA", id_ag, e)
        return {"ok": False, "id_agente": id_ag, "erro": "falha_pagina1"}

    if not visitas:
        return {"ok": False, "id_agente": id_ag, "nome_agente": nome}

    df = calcular_tempos(visitas, nome, id_ag, "Ponto Estratégico", semana_html_id=cfg_pe["semana_inicio"])

    semanas = []
    for semana_num, sub in df.groupby("semana", observed=True):
        if sub.empty:
            continue
        resumo_semana = resumo_agente(sub)
        chegada_min = sub["dt_chegada"].min()
        chegada_max = sub["dt_chegada"].max()
        periodo = (f"{chegada_min.strftime('%d/%m')} a {chegada_max.strftime('%d/%m')}"
                   if pd.notna(chegada_min) and pd.notna(chegada_max) else "")
        semanas.append({
            "nome_agente": nome,
            "equipe": "Ponto Estratégico",
            "semana": int(semana_num),
            "periodo_semana": periodo,
            "resumo": resumo_semana,
        })

    return {
        "ok": True,
        "id_agente": id_ag,
        "nome_agente": nome,
        "df": df,
        "paginas_falhas": paginas_falhas,
        "semanas": semanas,
    }


def _preparar_ponto_estrategico(semanas_todos_agentes):
    registros = []
    for item in semanas_todos_agentes:
        r, classif, fill = _formatar_campos_resumo(item["resumo"])
        registros.append({
            **r,
            "nome_agente": item["nome_agente"],
            "equipe": item["equipe"],
            "semana": item["semana"],
            "periodo_semana": item["periodo_semana"],
            "classificacao": classif,
            "_fill": fill,
        })
    registros.sort(key=lambda x: (x["nome_agente"], x["semana"]))
    return registros


# ============================================================
# CACHE POR SEMANA (data/semanas/) — coleta e consolidação
# ============================================================
# Cada semana vira UMA planilha própria (visitas + colunas calculadas) mais
# um JSON pequeno com os números de pendência (fechados/recusados) daquela
# semana. Consolidar N semanas é: carregar as N planilhas, juntar tudo e
# recalcular os resumos por agente em cima do conjunto combinado — os
# mesmos cálculos de sempre (resumo_agente etc.), só que agora alimentados
# pelo cache em vez de uma coleta ao vivo.

_COLS_BOOL_SEMANA = [
    "alerta_rapida", "alerta_longa", "acima_teto_media", "alerta_neg",
    "dup_por_id", "dup_por_chegada_imovel", "duplicada",
    "par_suspeito_anterior", "sequencia_suspeita",
    "fim_de_semana", "dentro_expediente", "fora_expediente",
    "visita_horario_almoco",
]
_COLS_DATETIME_SEMANA = ["dt_chegada", "dt_saida"]


def _caminho_semana(html_id, sufixo="", pasta_semanas=PASTA_SEMANAS):
    nome = f"{html_id}{sufixo}"
    return os.path.join(pasta_semanas, nome)


def _salvar_df_semana(df, caminho_xlsx):
    os.makedirs(os.path.dirname(caminho_xlsx), exist_ok=True)
    df.to_excel(caminho_xlsx, sheet_name="visitas", index=False)


def _carregar_df_semana(caminho_xlsx):
    """Recarrega uma planilha de semana cacheada, corrigindo os tipos que o
    Excel não guarda sozinho (datas viram texto/serial, booleanos podem
    virar 'True'/'False' em string etc.)."""
    if not os.path.exists(caminho_xlsx):
        return pd.DataFrame()
    df = pd.read_excel(caminho_xlsx, sheet_name="visitas")
    for col in _COLS_DATETIME_SEMANA:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce")
    if "dia" in df.columns:
        df["dia"] = pd.to_datetime(df["dia"], errors="coerce").dt.date
    for col in _COLS_BOOL_SEMANA:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip().str.lower().map(
                {"true": True, "1": True, "1.0": True,
                 "false": False, "0": False, "0.0": False}
            ).fillna(False)
    return df


def salvar_cache_semana(html_id, df_tratamento, pendencia_por_agente, df_pe=None):
    """Grava em disco tudo que foi coletado para UMA semana: a planilha de
    visitas (tratamento), a planilha de Ponto Estratégico (se houve) e um
    JSON com os números de pendência (fechados/recusados) por agente."""
    os.makedirs(PASTA_SEMANAS, exist_ok=True)
    if df_tratamento is not None and not df_tratamento.empty:
        _salvar_df_semana(df_tratamento, _caminho_semana(html_id, ".xlsx"))
    with open(_caminho_semana(html_id, "_pendencia.json"), "w", encoding="utf-8") as f:
        json.dump(pendencia_por_agente, f, ensure_ascii=False, indent=2)
    if df_pe is not None and not df_pe.empty:
        _salvar_df_semana(df_pe, _caminho_semana(html_id, "_pe.xlsx"))
    log.info("💾 Semana %s (Semana do Ano %s) cacheada em %s",
              html_id, sem.html_para_semana(html_id, sem.ano_da_semana_html(html_id)), PASTA_SEMANAS)


def coletar_uma_semana(session, html_id, id_ano, id_ciclo):
    """Coleta TODOS os agentes (tratamento + Ponto Estratégico, quando
    aplicável) para uma única semana (HTML ID) e já salva o cache em disco
    antes de retornar. Retorna um dict com estatísticas de integridade pra
    logar no resumo final."""
    id_ciclo_semana = ciclo_tratamento_para_semana(html_id, id_ciclo)
    cfg = {"semana_inicio": html_id, "semana_fim": html_id, "id_ano": id_ano, "id_ciclo": id_ciclo_semana}
    trabalhos = [(eq, ag) for eq, agentes in AGENTES_POR_EQUIPE.items() for ag in agentes]
    log.info("🚀 Semana %s (%s) | %s agentes | ano %s | ciclo %s",
              html_id, sem.rotulo_semana_html(html_id), len(trabalhos), id_ano, id_ciclo_semana)

    frames, pendencia_por_agente = [], {}
    sessao_caiu = False
    agentes_ausentes, agentes_incompletos = [], []

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as ex:
        futuros = {ex.submit(processar, eq, ag, session, cfg): (eq, ag) for eq, ag in trabalhos}
        for fut in as_completed(futuros):
            eq, ag = futuros[fut]
            try:
                ret = fut.result()
                if ret["ok"]:
                    frames.append(ret["df"])
                    pendencia_por_agente[str(ret["id_agente"])] = {
                        "nome_agente": ret["nome_agente"],
                        "equipe": ret["equipe"],
                        "imoveis_fechados": ret["resumo"]["imoveis_fechados"],
                        "imoveis_recusados": ret["resumo"]["imoveis_recusados"],
                        "fechados_por_area": ret["fechados_por_area"],
                        "recusados_por_area": ret["recusados_por_area"],
                        "visitas_recusados_raw": ret["visitas_recusados_raw"],
                        "paginas_falhas": ret["paginas_falhas"],
                    }
                    r = ret["resumo"]
                    log.info("✅ %s | %s visitas | Duplicadas: %s",
                              ret["nome_agente"], r["total_visitas"], r["visitas_duplicadas"])
                    if ret.get("paginas_falhas"):
                        agentes_incompletos.append((ret["nome_agente"], eq, ret["paginas_falhas"]))
                elif ret.get("erro") == "sessao_expirada":
                    sessao_caiu = True
                    agentes_ausentes.append((ag, eq, "sessão expirou durante a coleta"))
                elif ret.get("erro") == "falha_pagina1":
                    agentes_ausentes.append((ag, eq, "falha ao coletar página 1 (mesmo após retries)"))
                else:
                    log.warning("⚠️ Sem visitas para agente %s (%s) na semana %s", ag, eq, html_id)
            except Exception as e:
                log.error("❌ Erro processando agente %s (%s) na semana %s: %s", ag, eq, html_id, e)
                agentes_ausentes.append((ag, eq, f"erro inesperado: {e}"))

    if sessao_caiu:
        log.error("A sessão expirou durante a coleta da semana %s — refaça o login e rode de novo.", html_id)

    df_tratamento = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()

    # Ponto Estratégico — só busca se houver ciclo cadastrado pra essa semana.
    df_pe = pd.DataFrame()
    id_ciclo_pe = ciclo_para_semana(html_id)
    if AGENTES_PONTO_ESTRATEGICO and id_ciclo_pe is None:
        log.warning("⚠️ Ponto Estratégico: semana %s cai num ciclo sem id_ciclo cadastrado em "
                     "CICLOS_PONTO_ESTRATEGICO — pulando (confira o id_ciclo no site e cadastre-o).", html_id)
    elif AGENTES_PONTO_ESTRATEGICO:
        cfg_pe = {"semana_inicio": html_id, "semana_fim": html_id, "id_ano": id_ano, "id_ciclo": id_ciclo_pe}
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as ex:
            futuros_pe = {ex.submit(processar_ponto_estrategico, ag, session, cfg_pe): ag
                          for ag in AGENTES_PONTO_ESTRATEGICO}
            frames_pe = []
            for fut in as_completed(futuros_pe):
                ag = futuros_pe[fut]
                try:
                    ret = fut.result()
                    if ret["ok"]:
                        frames_pe.append(ret["df"])
                        log.info("✅ Ponto Estratégico — Agente %s (semana %s)", ret["nome_agente"], html_id)
                    else:
                        log.warning("⚠️ Ponto Estratégico — sem dados pro agente %s na semana %s (%s)",
                                     ag, html_id, ret.get("erro", "sem visitas"))
                except Exception as e:
                    log.error("❌ Erro no Ponto Estratégico, agente %s, semana %s: %s", ag, html_id, e)
        if frames_pe:
            df_pe = pd.concat(frames_pe, ignore_index=True)

    salvar_cache_semana(html_id, df_tratamento, pendencia_por_agente, df_pe)

    return {
        "html_id": html_id, "sessao_caiu": sessao_caiu,
        "agentes_ausentes": agentes_ausentes, "agentes_incompletos": agentes_incompletos,
        "total_agentes": len(trabalhos),
    }


def _salvar_agentes_nomes(resultados, caminho="data/agentes_nomes.json"):
    """Salva o mapa ID -> Nome de cada agente (id_agente, nome_agente,
    equipe) toda vez que consolida — é a única fonte confiável desse mapa,
    já que os nomes só existem de verdade dentro do e-Visita (o
    AGENTES_POR_EQUIPE deste script só tem os IDs). Serve pra
    cronograma_editor.html mostrar nomes de verdade no menu em vez de só
    "ID 116", sem precisar digitar nada — é só colar esse arquivo lá."""
    mapa = {str(r["id_agente"]): {"nome": r["nome_agente"], "equipe": r["equipe"]} for r in resultados}
    try:
        with open(caminho, "w", encoding="utf-8") as f:
            json.dump(mapa, f, ensure_ascii=False, indent=2, sort_keys=True)
        log.info("💾 Nomes dos agentes salvos em %s (%s agentes).", caminho, len(mapa))
    except OSError as e:
        log.warning("⚠️ Não deu pra salvar %s: %s", caminho, e)


def consolidar_semanas(htmls, pasta_saida):
    """Carrega o cache de cada semana em `htmls`, junta tudo e recalcula os
    resumos por agente (e por semana, no caso do Ponto Estratégico) em cima
    do conjunto combinado. Gera o Analise_Consolidada.xlsx com a MESMA
    função de sempre (salvar_excel_consolidado), só que alimentada pelo
    cache em vez de uma coleta ao vivo — o formato do Excel final não muda
    em nada."""
    htmls = sorted(set(htmls))
    frames, frames_pe = [], []
    pendencia = {}   # id_agente (str) -> lista de dicts de pendência, uma por semana
    for h in htmls:
        df_h = _carregar_df_semana(_caminho_semana(h, ".xlsx"))
        if not df_h.empty:
            frames.append(df_h)
        df_pe_h = _carregar_df_semana(_caminho_semana(h, "_pe.xlsx"))
        if not df_pe_h.empty:
            frames_pe.append(df_pe_h)
        pend_path = _caminho_semana(h, "_pendencia.json")
        if os.path.exists(pend_path):
            with open(pend_path, encoding="utf-8") as f:
                for id_ag, info in json.load(f).items():
                    pendencia.setdefault(id_ag, []).append(info)

    if not frames:
        log.error("Nenhuma semana com dados encontrada no cache (%s) para consolidar.", htmls)
        return None

    df_total = pd.concat(frames, ignore_index=True)

    resultados = []
    for id_ag, sub in df_total.groupby("id_agente"):
        sub = sub.sort_values("dt_chegada").reset_index(drop=True)
        nome = sub["nome_agente"].iloc[0]
        equipe = sub["equipe"].iloc[0]
        resumo = resumo_agente(sub)

        pend_semanas = pendencia.get(str(id_ag), [])
        total_fechados = sum(p["imoveis_fechados"] for p in pend_semanas)
        total_recusados = sum(p["imoveis_recusados"] for p in pend_semanas)
        total_abertos = resumo["total_visitas"]
        total_geral_pendencia = total_abertos + total_fechados + total_recusados
        resumo["imoveis_fechados"] = total_fechados
        resumo["imoveis_recusados"] = total_recusados
        resumo["total_geral_pendencia"] = total_geral_pendencia
        resumo["pct_pendencia"] = round(((total_fechados + total_recusados) / total_geral_pendencia * 100), 2) \
            if total_geral_pendencia > 0 else 0

        fechados_por_area, recusados_por_area, visitas_recusados_raw, paginas_falhas = {}, {}, [], []
        for p in pend_semanas:
            for area, qtd in p.get("fechados_por_area", {}).items():
                fechados_por_area[area] = fechados_por_area.get(area, 0) + qtd
            for area, qtd in p.get("recusados_por_area", {}).items():
                recusados_por_area[area] = recusados_por_area.get(area, 0) + qtd
            visitas_recusados_raw.extend(p.get("visitas_recusados_raw", []))
            paginas_falhas.extend(p.get("paginas_falhas", []))

        resultados.append({
            "ok": True, "id_agente": id_ag, "nome_agente": nome, "equipe": equipe,
            "resumo": resumo, "df": sub, "paginas_falhas": paginas_falhas,
            "fechados_por_area": fechados_por_area, "recusados_por_area": recusados_por_area,
            "visitas_recusados_raw": visitas_recusados_raw,
        })
    resultados.sort(key=lambda r: (r["equipe"], r["nome_agente"]))

    _salvar_agentes_nomes(resultados)

    semanas_ponto_estrategico = []
    if frames_pe:
        df_pe_total = pd.concat(frames_pe, ignore_index=True)
        for (id_ag, semana_num), sub in df_pe_total.groupby(["id_agente", "semana"], observed=True):
            if sub.empty:
                continue
            nome = sub["nome_agente"].iloc[0]
            resumo_semana = resumo_agente(sub)
            chegada_min, chegada_max = sub["dt_chegada"].min(), sub["dt_chegada"].max()
            periodo = (f"{chegada_min.strftime('%d/%m')} a {chegada_max.strftime('%d/%m')}"
                       if pd.notna(chegada_min) and pd.notna(chegada_max) else "")
            semanas_ponto_estrategico.append({
                "nome_agente": nome, "equipe": "Ponto Estratégico", "semana": int(semana_num),
                "periodo_semana": periodo, "resumo": resumo_semana,
            })

    # Resumo Semanal (tratamento) — mesma quebra por semana, só que pros
    # agentes normais, pra alimentar o filtro de semana do dashboard. Só
    # vale a pena gerar se o período consolidado cobrir mais de 1 semana.
    semanas_tratamento = []
    if len(htmls) > 1:
        for (id_ag, semana_num), sub in df_total.groupby(["id_agente", "semana"], observed=True):
            if sub.empty:
                continue
            nome = sub["nome_agente"].iloc[0]
            equipe = sub["equipe"].iloc[0]
            resumo_semana = resumo_agente(sub)
            chegada_min, chegada_max = sub["dt_chegada"].min(), sub["dt_chegada"].max()
            periodo = (f"{chegada_min.strftime('%d/%m')} a {chegada_max.strftime('%d/%m')}"
                       if pd.notna(chegada_min) and pd.notna(chegada_max) else "")
            semanas_tratamento.append({
                "nome_agente": nome, "equipe": equipe, "semana": int(semana_num),
                "periodo_semana": periodo, "resumo": resumo_semana,
            })

    caminho = salvar_excel_consolidado(resultados, df_total, pasta_saida,
                                        semanas_ponto_estrategico, semanas_tratamento)
    log.info("🎉 Consolidado %s semana(s) (HTML %s-%s) em: %s", len(htmls), htmls[0], htmls[-1], caminho)
    return caminho


def atualizar_historico_do_cache(pasta_semanas=PASTA_SEMANAS, pasta_historico="data/historico",
                                  pasta_saida_tmp="output/_tmp_historico"):
    """Reconstrói as entradas AUTOMÁTICAS de data/historico/ a partir do cache
    de semanas: agrupa as semanas já coletadas por mês civil e gera (ou
    regera) uma planilha por mês. Cada entrada do manifest.json tem um campo
    "origem": "cache" (escrita por esta função, pode ser sobrescrita à
    vontade) ou "manual" (adicionada por fora — ex.: uma planilha antiga de
    Jan-Maio/2026 coletada com outros parâmetros — e NUNCA tocada aqui,
    mesmo que não tenha nenhuma semana correspondente no cache).

    Pra adicionar uma entrada manual: coloque o arquivo .xlsx em
    data/historico/, e acrescente uma entrada em data/historico/manifest.json
    com "origem": "manual" (ver README, seção "Planilha antiga / fora do
    sistema de coleta"). Essa entrada nunca some, mesmo que o cache de
    semanas mude — e como ela nunca passa por consolidar_semanas(), nunca
    entra em data/Analise_Consolidada.xlsx nem em acumulado.html."""
    htmls = semanas_ja_coletadas(pasta_semanas)

    manifest_path = os.path.join(pasta_historico, "manifest.json")
    manifest_anterior = []
    if os.path.exists(manifest_path):
        try:
            manifest_anterior = json.loads(open(manifest_path, encoding="utf-8").read())
        except (json.JSONDecodeError, OSError):
            manifest_anterior = []
    entradas_manuais = [e for e in manifest_anterior if e.get("origem") == "manual"]

    if not htmls:
        log.warning("Nenhuma semana em cache — mantendo só as entradas manuais do histórico (se houver).")
        with open(manifest_path, "w", encoding="utf-8") as f:
            json.dump(entradas_manuais, f, ensure_ascii=False, indent=2)
        return

    por_mes = {}
    for h in htmls:
        por_mes.setdefault(sem.mes_da_semana_html(h), []).append(h)

    os.makedirs(pasta_historico, exist_ok=True)
    os.makedirs(pasta_saida_tmp, exist_ok=True)
    entradas_cache = []
    for (ano, mes) in sorted(por_mes):
        semanas_mes = sorted(por_mes[(ano, mes)])
        ini, fim = semanas_mes[0], semanas_mes[-1]
        nome_arquivo = f"semanas_{ini}-{fim}.xlsx" if ini != fim else f"semana_{ini}.xlsx"
        label = nome_periodo(ini, fim)

        caminho_tmp = consolidar_semanas(semanas_mes, pasta_saida_tmp)
        if not caminho_tmp:
            continue
        destino = os.path.join(pasta_historico, nome_arquivo)
        shutil.copy2(caminho_tmp, destino)

        entradas_cache.append({
            "arquivo": nome_arquivo, "label": label, "origem": "cache",
            "semana_inicio": ini, "semana_fim": fim,
            "semana_ano_inicio": sem.html_para_semana(ini, sem.ano_da_semana_html(ini)),
            "semana_ano_fim": sem.html_para_semana(fim, sem.ano_da_semana_html(fim)),
            "gerado_em": datetime.now().strftime("%d/%m/%Y %H:%M"),
        })
        log.info("🗂️  Histórico atualizado: %s (%s)", destino, label)

    # Entradas manuais primeiro (geralmente mais antigas, tipo Jan-Maio),
    # depois as automáticas em ordem cronológica — mas isso é só estético,
    # a ordem não afeta nada funcionalmente.
    manifest = entradas_manuais + entradas_cache
    with open(manifest_path, "w", encoding="utf-8") as f:
        json.dump(manifest, f, ensure_ascii=False, indent=2)
    shutil.rmtree(pasta_saida_tmp, ignore_errors=True)
    if entradas_manuais:
        log.info("🔒 %s entrada(s) manual(is) do histórico preservada(s): %s",
                  len(entradas_manuais), [e.get("label") for e in entradas_manuais])


def main():
    args = parse_args()

    if args.semana is not None:
        semanas_a_coletar = [args.semana]
        log.info("▶️  Modo manual: recoletando só a semana %s (HTML ID).", args.semana)
    else:
        semanas_a_coletar = decidir_semanas_a_coletar(pasta_semanas=args.pasta_semanas)

    if not semanas_a_coletar:
        log.info("Nada novo para coletar — semana atual já está em cache.")
    else:
        log.info("📋 Semanas a coletar nesta execução: %s", semanas_a_coletar)

        driver = criar_driver()
        try:
            fazer_login(driver)
            session = criar_session(driver)
        finally:
            driver.quit()

        resumo_integridade = []
        for html_id in semanas_a_coletar:
            resumo_integridade.append(coletar_uma_semana(session, html_id, args.id_ano, args.id_ciclo))

        ausentes_total = sum(len(r["agentes_ausentes"]) for r in resumo_integridade)
        incompletos_total = sum(len(r["agentes_incompletos"]) for r in resumo_integridade)
        if ausentes_total or incompletos_total:
            log.error("=" * 60)
            log.error("⚠️  RESUMO DE INTEGRIDADE DA COLETA — LEIA COM ATENÇÃO")
            for r in resumo_integridade:
                for id_ag, eq, motivo in r["agentes_ausentes"]:
                    log.error("  - [Semana %s] Agente %s (%s): %s", r["html_id"], id_ag, eq, motivo)
                for nome, eq, paginas in r["agentes_incompletos"]:
                    log.error("  - [Semana %s] %s (%s): páginas %s não coletadas",
                               r["html_id"], nome, eq, paginas)
            log.error("Recomendação: rode `python scripts/coletar_evisita.py --semana <HTML_ID>` "
                       "pra recoletar só a(s) semana(s) afetada(s).")
            log.error("=" * 60)
        else:
            log.info("✅ Coleta íntegra em todas as semanas processadas.")

    # Consolidação: sempre reconstrói o Analise_Consolidada.xlsx a partir de
    # TODO o cache disponível (não só das semanas coletadas agora), e
    # reconstrói o histórico por mês em cima do mesmo cache.
    todas_semanas = semanas_ja_coletadas(args.pasta_semanas if args.semana is None else PASTA_SEMANAS)
    if not todas_semanas:
        log.error("Nenhuma semana em cache — não há o que consolidar.")
        return

    caminho = consolidar_semanas(todas_semanas, args.saida)
    if caminho:
        log.info("🎉 CONCLUÍDO! Analise_Consolidada.xlsx gerado a partir de %s semana(s) em cache.",
                   len(todas_semanas))
    atualizar_historico_do_cache(pasta_semanas=args.pasta_semanas)


if __name__ == "__main__":
    main()
