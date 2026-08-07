import openpyxl
import pandas as pd

def scan_tables(ws, max_col=32):
    tables = []
    current_header = None
    current_data = []
    def finalize():
        nonlocal current_header, current_data
        if current_header is not None and current_data:
            tables.append((current_header, current_data))
        current_header = None
        current_data = []
    for r in range(1, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, max_col + 1)]
        is_header = ('Agente' in vals) and ('Equipe' in vals)
        non_null = sum(v is not None for v in vals)
        first = vals[0]
        is_total = isinstance(first, str) and first.strip().upper().startswith('TOTAL')
        is_title_or_blank = non_null <= 1
        if is_header:
            finalize()
            current_header = vals
            continue
        if is_title_or_blank or is_total:
            finalize()
            continue
        if current_header is not None:
            current_data.append(vals)
    finalize()
    return tables

def tables_to_dfs(tables):
    dfs = []
    for header, data in tables:
        n = len(header)
        while n > 0 and header[n-1] is None:
            n -= 1
        header = header[:n]
        rows = [row[:n] for row in data]
        df = pd.DataFrame(rows, columns=header)
        dfs.append(df)
    return dfs

def get_tables(path, sheet_name):
    wb = openpyxl.load_workbook(path, data_only=True)
    return tables_to_dfs(scan_tables(wb[sheet_name]))

def get_table(path, sheet_name, index=0):
    return get_tables(path, sheet_name)[index]
